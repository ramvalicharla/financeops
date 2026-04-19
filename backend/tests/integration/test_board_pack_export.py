from __future__ import annotations

import sys
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from financeops.db.models.board_pack_generator import BoardPackGeneratorArtifact
from financeops.db.rls import set_tenant_context
from financeops.modules.board_pack_generator.application.generate_service import (
    BoardPackGenerateService,
)
from financeops.modules.board_pack_generator.application.export_service import (
    BoardPackExportService,
)
from financeops.modules.board_pack_generator.domain.enums import PeriodType, SectionType
from financeops.modules.board_pack_generator.domain.pack_definition import (
    AssembledPack,
    PackDefinitionSchema,
    RenderedSection,
    SectionConfig,
)
from financeops.modules.board_pack_generator.infrastructure.repository import (
    BoardPackRepository,
)
from tests.integration.test_board_pack_generate_service import _relax_board_pack_schema


class _FakeStorage:
    def __init__(self) -> None:
        self.uploads: list[dict[str, object]] = []
        self.signed: list[tuple[str, int]] = []

    def upload_file(
        self,
        file_bytes: bytes,
        key: str,
        content_type: str,
        tenant_id: str,
        uploaded_by: str | None = None,
    ) -> str:
        self.uploads.append(
            {
                "bytes": file_bytes,
                "key": key,
                "content_type": content_type,
                "tenant_id": tenant_id,
                "uploaded_by": uploaded_by,
            }
        )
        return key

    def generate_signed_url(self, key: str, expires_in: int = 3600) -> str:
        self.signed.append((key, expires_in))
        return f"https://signed.example.com/{key}?exp={expires_in}&sig=test"


def _make_pack() -> AssembledPack:
    section = RenderedSection(
        section_type=SectionType.PROFIT_AND_LOSS,
        section_order=1,
        title="P&L",
        data_snapshot={"revenue": Decimal("1000000.00"), "variance": Decimal("10.00")},
        section_hash="hash-1",
    )
    return AssembledPack(
        run_id=uuid.uuid4(),
        tenant_id=uuid.uuid4(),
        period_start=date(2026, 1, 1),
        period_end=date(2026, 1, 31),
        sections=[section],
        chain_hash="chain-hash",
    )


async def _seed_definition_and_run(
    session: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    user_id: uuid.UUID,
) -> uuid.UUID:
    await _relax_board_pack_schema(session)
    repo = BoardPackRepository()
    schema = PackDefinitionSchema(
        name="Board Pack",
        description="Monthly board pack",
        section_configs=[
            SectionConfig(section_type=SectionType.PROFIT_AND_LOSS, order=1),
        ],
        entity_ids=[uuid.uuid4()],
        period_type=PeriodType.MONTHLY,
        config={},
    )
    definition = await repo.create_definition(
        db=session,
        tenant_id=tenant_id,
        schema=schema,
        created_by=user_id,
    )
    run = await repo.create_run(
        db=session,
        tenant_id=tenant_id,
        definition_id=definition.id,
        period_start=date(2026, 1, 1),
        period_end=date(2026, 1, 31),
        triggered_by=user_id,
    )
    await session.commit()
    await set_tenant_context(session, tenant_id)
    return run.id


def _install_fake_weasyprint(monkeypatch: pytest.MonkeyPatch) -> None:
    class FakeHTML:
        def __init__(self, string: str, base_url: str | None = None) -> None:
            self.string = string
            self.base_url = base_url

        def write_pdf(self) -> bytes:
            return b"%PDF-1.7\nboard-pack\n"

    monkeypatch.setitem(sys.modules, "weasyprint", SimpleNamespace(HTML=FakeHTML))


@pytest.mark.integration
def test_pdf_export_returns_valid_pdf_bytes_not_empty(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from financeops.modules.board_pack_generator.application import export_service

    _install_fake_weasyprint(monkeypatch)
    monkeypatch.setattr(export_service, "assert_weasyprint_available", lambda: None)

    payload, filename = BoardPackExportService().export_pdf(
        _make_pack(),
        "Pack",
        datetime.now(UTC),
    )

    assert filename.endswith(".pdf")
    assert payload.startswith(b"%PDF")
    assert len(payload) > 10


@pytest.mark.integration
def test_excel_export_returns_valid_xlsx_bytes() -> None:
    payload, filename = BoardPackExportService().export_excel(
        _make_pack(),
        "Pack",
        datetime.now(UTC),
    )

    workbook = load_workbook(BytesIO(payload))
    assert filename.endswith(".xlsx")
    assert workbook.sheetnames[0] == "Cover"
    assert "P&L" in workbook.sheetnames


@pytest.mark.integration
@pytest.mark.asyncio
async def test_export_artifact_inserted_as_append_only_row(
    async_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fake_storage = _FakeStorage()
    _install_fake_weasyprint(monkeypatch)
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.application.export_service.assert_weasyprint_available",
        lambda: None,
    )
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.application.generate_service.get_storage",
        lambda: fake_storage,
    )

    await set_tenant_context(async_session, test_user.tenant_id)
    run_id = await _seed_definition_and_run(
        async_session,
        tenant_id=test_user.tenant_id,
        user_id=test_user.id,
    )
    service = BoardPackGenerateService()

    async def _fake_fetch(*, db, context, section_type):  # noqa: ANN001
        return {"section": section_type.value, "amount": Decimal("10.00")}

    service._fetch_section_source_data = _fake_fetch  # type: ignore[method-assign]
    await service.generate(db=async_session, run_id=run_id, tenant_id=test_user.tenant_id)
    latest_run_id = (
        await async_session.execute(
            select(BoardPackGeneratorArtifact.run_id)
            .where(BoardPackGeneratorArtifact.tenant_id == test_user.tenant_id)
            .order_by(BoardPackGeneratorArtifact.generated_at.desc(), BoardPackGeneratorArtifact.id.desc())
            .limit(1)
        )
    ).scalar_one()
    await service.export_run_artifacts(
        db=async_session,
        run_id=latest_run_id,
        tenant_id=test_user.tenant_id,
    )
    await async_session.commit()

    artifacts = list(
        (
            await async_session.execute(
                select(BoardPackGeneratorArtifact)
                .where(BoardPackGeneratorArtifact.tenant_id == test_user.tenant_id)
                .order_by(BoardPackGeneratorArtifact.generated_at.asc(), BoardPackGeneratorArtifact.id.asc())
            )
        ).scalars()
    )

    assert len(artifacts) == 4
    assert len({artifact.id for artifact in artifacts}) == 4
    assert len({artifact.storage_path for artifact in artifacts}) == 4


@pytest.mark.integration
@pytest.mark.asyncio
async def test_artifact_uploaded_to_r2_not_local_disk(
    async_session: AsyncSession,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fake_storage = _FakeStorage()
    _install_fake_weasyprint(monkeypatch)
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.application.export_service.assert_weasyprint_available",
        lambda: None,
    )
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.application.generate_service.get_storage",
        lambda: fake_storage,
    )
    monkeypatch.setattr(
        Path,
        "write_bytes",
        lambda self, data: (_ for _ in ()).throw(AssertionError("local disk writes forbidden")),
    )

    await set_tenant_context(async_session, test_user.tenant_id)
    run_id = await _seed_definition_and_run(
        async_session,
        tenant_id=test_user.tenant_id,
        user_id=test_user.id,
    )
    service = BoardPackGenerateService()

    async def _fake_fetch(*, db, context, section_type):  # noqa: ANN001
        return {"section": section_type.value, "amount": Decimal("10.00")}

    service._fetch_section_source_data = _fake_fetch  # type: ignore[method-assign]
    await service.generate(db=async_session, run_id=run_id, tenant_id=test_user.tenant_id)

    assert len(fake_storage.uploads) == 2
    assert all(upload["key"] for upload in fake_storage.uploads)
    assert all("artifacts/board_packs/" in str(upload["key"]) for upload in fake_storage.uploads)


@pytest.mark.integration
@pytest.mark.asyncio
async def test_download_url_is_signed_expires_in_15_min_not_public(
    async_client,
    test_access_token: str,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    fake_storage = _FakeStorage()
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.api.routes.get_storage",
        lambda: fake_storage,
    )
    run_id = uuid.uuid4()
    artifact = BoardPackGeneratorArtifact(
        id=uuid.uuid4(),
        run_id=run_id,
        tenant_id=test_user.tenant_id,
        format="PDF",
        storage_path=f"artifacts/board_packs/{test_user.tenant_id}/{run_id}/x/board_pack.pdf",
        file_size_bytes=128,
        checksum="a" * 64,
    )

    async def _fake_get_run(self, db, tenant_id, run_id):  # noqa: ANN001
        del self, db
        if tenant_id != test_user.tenant_id:
            return None
        return SimpleNamespace(id=run_id, tenant_id=tenant_id)

    async def _fake_get_artifact(self, db, tenant_id, artifact_id):  # noqa: ANN001
        del self, db
        if tenant_id != test_user.tenant_id or artifact_id != artifact.id:
            return None
        return artifact

    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.api.routes.BoardPackRepository.get_run",
        _fake_get_run,
    )
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.api.routes.BoardPackRepository.get_artifact",
        _fake_get_artifact,
    )

    response = await async_client.get(
        f"/api/v1/board-packs/runs/{run_id}/export/{artifact.id}",
        headers={"Authorization": f"Bearer {test_access_token}"},
    )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["artifact_id"] == str(artifact.id)
    assert payload["expires_in_seconds"] == 900
    assert payload["signed_url"].startswith("https://signed.example.com/")
    assert "public" not in payload["signed_url"]
    assert fake_storage.signed == [(artifact.storage_path, 900)]


@pytest.mark.integration
@pytest.mark.asyncio
async def test_export_route_returns_202_not_200(
    async_client,
    test_access_token: str,
    test_user,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    run_id = uuid.uuid4()

    async def _fake_get_run(self, db, tenant_id, run_id):  # noqa: ANN001
        del self, db
        if tenant_id != test_user.tenant_id:
            return None
        return SimpleNamespace(id=run_id, tenant_id=tenant_id)

    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.api.routes.BoardPackRepository.get_run",
        _fake_get_run,
    )
    monkeypatch.setattr(
        "financeops.modules.board_pack_generator.api.routes.export_board_pack_artifacts_task.delay",
        lambda run_id, tenant_id: SimpleNamespace(id="task-export-123"),
    )

    response = await async_client.post(
        f"/api/v1/board-packs/runs/{run_id}/export",
        headers={"Authorization": f"Bearer {test_access_token}"},
    )

    assert response.status_code == 202
    payload = response.json()["data"]
    assert payload["task_id"] == "task-export-123"
    assert payload["status"] == "queued"
