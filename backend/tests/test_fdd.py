from __future__ import annotations

import io
import uuid
from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from typing import Any

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from financeops.core.exceptions import InsufficientCreditsError
from financeops.db.append_only import append_only_function_sql, create_trigger_sql, drop_trigger_sql
from financeops.db.models.credits import CreditReservation, ReservationStatus
from financeops.db.models.users import IamUser
from financeops.modules.fdd.models import FDDEngagement, FDDFinding, FDDSection
from financeops.modules.fdd.sections import debt_liability_review, quality_of_earnings, working_capital_analysis
from financeops.modules.fdd.sections.headcount_normalisation import compute_headcount_normalisation
from financeops.modules.fdd.sections.quality_of_earnings import compute_qoe
from financeops.modules.fdd.sections.revenue_quality import compute_revenue_quality
from financeops.modules.fdd.sections.working_capital_analysis import compute_wc_analysis
from financeops.modules.fdd.service import (
    create_engagement,
    export_fdd_report,
    get_engagement_report,
    run_engagement,
)
from financeops.modules.working_capital.models import WCSnapshot
from financeops.services.credit_service import add_credits


async def _fund(async_session: AsyncSession, tenant_id: uuid.UUID, amount: str = "10000.00") -> None:
    await add_credits(async_session, tenant_id, Decimal(amount), "test_fdd_fund")
    await async_session.flush()


async def _create_default_engagement(async_session: AsyncSession, user: IamUser) -> FDDEngagement:
    return await create_engagement(
        async_session,
        tenant_id=user.tenant_id,
        engagement_name="Acme Corp FDD",
        target_company_name="Acme Corp",
        analysis_period_start=date(2024, 1, 1),
        analysis_period_end=date(2024, 12, 31),
        sections_requested=[
            "quality_of_earnings",
            "working_capital",
            "debt_liability",
            "headcount",
            "revenue_quality",
        ],
        created_by=user.id,
    )


async def _seed_wc_snapshot(
    async_session: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    period: str,
    nwc: Decimal,
    dso: Decimal = Decimal("10.00"),
    dpo: Decimal = Decimal("8.00"),
) -> None:
    year, month = map(int, period.split("-"))
    async_session.add(
        WCSnapshot(
            tenant_id=tenant_id,
            period=period,
            entity_id=None,
            snapshot_date=date(year, month, 28),
            ar_total=Decimal("100.00"),
            ar_current=Decimal("40.00"),
            ar_30=Decimal("30.00"),
            ar_60=Decimal("20.00"),
            ar_90=Decimal("10.00"),
            dso_days=dso,
            ap_total=Decimal("50.00"),
            ap_current=Decimal("20.00"),
            ap_30=Decimal("15.00"),
            ap_60=Decimal("10.00"),
            ap_90=Decimal("5.00"),
            dpo_days=dpo,
            inventory_days=Decimal("0.00"),
            ccc_days=dso - dpo,
            net_working_capital=nwc,
            current_ratio=Decimal("1.5000"),
            quick_ratio=Decimal("1.2000"),
        )
    )
    await async_session.flush()


def _contains_float(value: Any) -> bool:
    if isinstance(value, float):
        return True
    if isinstance(value, list):
        return any(_contains_float(item) for item in value)
    if isinstance(value, dict):
        return any(_contains_float(item) for item in value.values())
    return False


def _severity_rank(severity: str) -> int:
    order = {
        "critical": 0,
        "high": 1,
        "medium": 2,
        "low": 3,
        "informational": 4,
    }
    return order[severity]


# Engagement lifecycle (5)
@pytest.mark.asyncio
async def test_create_engagement_reserves_credits(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    reservation = (
        await async_session.execute(
            select(CreditReservation).where(
                CreditReservation.tenant_id == test_user.tenant_id,
                CreditReservation.task_type == f"fdd:{engagement.id}",
            )
        )
    ).scalar_one()
    assert engagement.status == "draft"
    assert reservation.status == ReservationStatus.pending
    assert engagement.credits_reserved_at is not None


@pytest.mark.asyncio
async def test_create_engagement_insufficient_credits(async_session: AsyncSession, test_user: IamUser) -> None:
    with pytest.raises(InsufficientCreditsError):
        await _create_default_engagement(async_session, test_user)


@pytest.mark.asyncio
async def test_run_engagement_completes_all_sections(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    for month in range(1, 13):
        await _seed_wc_snapshot(
            async_session,
            tenant_id=test_user.tenant_id,
            period=f"2024-{month:02d}",
            nwc=Decimal("100.00"),
        )
    engagement = await _create_default_engagement(async_session, test_user)
    updated = await run_engagement(async_session, test_user.tenant_id, engagement.id)
    assert updated.status == "completed"
    assert len(updated.sections_completed) == 5


@pytest.mark.asyncio
async def test_run_engagement_partial_section_failure(
    async_session: AsyncSession,
    test_user: IamUser,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from financeops.modules.fdd import service as fdd_service

    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)

    async def _fail(*args, **kwargs):  # noqa: ANN001, ANN002
        del args, kwargs
        raise RuntimeError("forced section failure")

    monkeypatch.setitem(fdd_service.SECTION_COMPUTERS, "debt_liability", _fail)
    updated = await run_engagement(async_session, test_user.tenant_id, engagement.id)
    rows = (
        await async_session.execute(
            select(FDDSection).where(
                FDDSection.engagement_id == engagement.id,
                FDDSection.tenant_id == test_user.tenant_id,
            )
        )
    ).scalars().all()
    by_name = {row.section_name: row for row in rows}
    assert updated.status == "completed"
    assert by_name["debt_liability"].status == "failed"
    assert any(row.status == "completed" for row in rows if row.section_name != "debt_liability")


@pytest.mark.asyncio
async def test_run_engagement_deducts_credits(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    reservation = (
        await async_session.execute(
            select(CreditReservation).where(
                CreditReservation.tenant_id == test_user.tenant_id,
                CreditReservation.task_type == f"fdd:{engagement.id}",
            )
        )
    ).scalar_one()
    refreshed = await async_session.get(FDDEngagement, engagement.id)
    assert reservation.status == ReservationStatus.confirmed
    assert refreshed.credits_deducted_at is not None


# Section computations (8)
@pytest.mark.asyncio
async def test_qoe_computes_adjusted_ebitda(
    async_session: AsyncSession,
    test_user: IamUser,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def _fake_rows(*args, **kwargs):  # noqa: ANN001, ANN002
        del args, kwargs
        return [
            {
                "period": "2025-01",
                "revenue": Decimal("100"),
                "cogs": Decimal("60"),
                "opex": Decimal("20"),
                "one_time": Decimal("5"),
                "owner_comp_addback": Decimal("0"),
                "recurring_revenue_pct": Decimal("0.70"),
                "top_5_customer_pct": Decimal("0.40"),
                "contract_coverage_pct": Decimal("0.75"),
            }
        ]

    monkeypatch.setattr(quality_of_earnings, "_load_qoe_period_data", _fake_rows)
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="qoe",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 1, 31),
        sections_requested=["quality_of_earnings"],
        created_by=test_user.id,
    )
    result = await compute_qoe(async_session, engagement)
    assert result["reported_ebitda"][0] == Decimal("20.00")
    assert result["adjusted_ebitda"][0] == Decimal("25.00")


@pytest.mark.asyncio
async def test_qoe_ebitda_cagr_is_decimal(async_session: AsyncSession, test_user: IamUser) -> None:
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="qoe",
        target_company_name="target",
        analysis_period_start=date(2024, 1, 1),
        analysis_period_end=date(2024, 12, 31),
        sections_requested=["quality_of_earnings"],
        created_by=test_user.id,
    )
    result = await compute_qoe(async_session, engagement)
    assert isinstance(result["ebitda_cagr"], Decimal)


@pytest.mark.asyncio
async def test_wc_analysis_nwc_peg_is_decimal(
    async_session: AsyncSession,
    test_user: IamUser,
) -> None:
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-01", nwc=Decimal("100.00"))
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-02", nwc=Decimal("100.00"))
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-03", nwc=Decimal("100.00"))
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="wc",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 3, 31),
        sections_requested=["working_capital"],
        created_by=test_user.id,
    )
    result = await compute_wc_analysis(async_session, engagement)
    assert isinstance(result["nwc_peg"], Decimal)


@pytest.mark.asyncio
async def test_wc_analysis_average_nwc_correct(
    async_session: AsyncSession,
    test_user: IamUser,
) -> None:
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-01", nwc=Decimal("100.00"))
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-02", nwc=Decimal("120.00"))
    await _seed_wc_snapshot(async_session, tenant_id=test_user.tenant_id, period="2025-03", nwc=Decimal("110.00"))
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="wc",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 3, 31),
        sections_requested=["working_capital"],
        created_by=test_user.id,
    )
    result = await compute_wc_analysis(async_session, engagement)
    assert result["average_nwc"] == Decimal("110.00")


@pytest.mark.asyncio
async def test_debt_review_net_debt_correct(
    async_session: AsyncSession,
    test_user: IamUser,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    async def _fake_debt_inputs(*args, **kwargs):  # noqa: ANN001, ANN002
        del args, kwargs
        return {
            "financial_debt": Decimal("500"),
            "cash_and_equivalents": Decimal("200"),
            "debt_like_items": [{"description": "x", "amount": Decimal("100"), "category": "other"}],
            "notes": [],
        }

    monkeypatch.setattr(debt_liability_review, "_load_debt_inputs", _fake_debt_inputs)
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="debt",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 1, 31),
        sections_requested=["debt_liability"],
        created_by=test_user.id,
    )
    result = await debt_liability_review.compute_debt_review(async_session, engagement)
    assert result["net_debt"] == Decimal("400.00")


@pytest.mark.asyncio
async def test_headcount_normalisation_returns_decimal(async_session: AsyncSession, test_user: IamUser) -> None:
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="hc",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 12, 31),
        sections_requested=["headcount"],
        created_by=test_user.id,
    )
    result = await compute_headcount_normalisation(async_session, engagement)
    assert isinstance(result["avg_cost_per_employee"], Decimal)


@pytest.mark.asyncio
async def test_revenue_quality_score_range(async_session: AsyncSession, test_user: IamUser) -> None:
    engagement = FDDEngagement(
        tenant_id=test_user.tenant_id,
        engagement_name="rev",
        target_company_name="target",
        analysis_period_start=date(2025, 1, 1),
        analysis_period_end=date(2025, 12, 31),
        sections_requested=["revenue_quality"],
        created_by=test_user.id,
    )
    result = await compute_revenue_quality(async_session, engagement)
    assert Decimal("0") <= result["quality_score"] <= Decimal("100")


@pytest.mark.asyncio
async def test_no_float_in_any_section_result(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    sections = (
        await async_session.execute(
            select(FDDSection).where(
                FDDSection.tenant_id == test_user.tenant_id,
                FDDSection.engagement_id == engagement.id,
            )
        )
    ).scalars().all()
    assert sections
    assert all(not _contains_float(section.result_data) for section in sections)


# Findings (5)
@pytest.mark.asyncio
async def test_findings_created_from_sections(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    findings = (
        await async_session.execute(
            select(FDDFinding).where(
                FDDFinding.tenant_id == test_user.tenant_id,
                FDDFinding.engagement_id == engagement.id,
            )
        )
    ).scalars().all()
    assert len(findings) >= 5


@pytest.mark.asyncio
async def test_finding_financial_impact_is_decimal(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    findings = (
        await async_session.execute(
            select(FDDFinding).where(
                FDDFinding.tenant_id == test_user.tenant_id,
                FDDFinding.engagement_id == engagement.id,
            )
        )
    ).scalars().all()
    assert findings
    assert all(isinstance(row.financial_impact, Decimal) or row.financial_impact is None for row in findings)


@pytest.mark.asyncio
async def test_findings_ordered_by_severity(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    report = await get_engagement_report(async_session, test_user.tenant_id, engagement.id)
    severities = [row.severity for row in report["findings"]]
    assert severities == sorted(severities, key=_severity_rank)


@pytest.mark.asyncio
async def test_fdd_findings_append_only(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    finding = (
        await async_session.execute(
            select(FDDFinding).where(
                FDDFinding.tenant_id == test_user.tenant_id,
                FDDFinding.engagement_id == engagement.id,
            )
        )
    ).scalars().first()
    assert finding is not None
    await async_session.execute(text(append_only_function_sql()))
    await async_session.execute(text(drop_trigger_sql("fdd_findings")))
    await async_session.execute(text(create_trigger_sql("fdd_findings")))
    await async_session.flush()
    with pytest.raises(Exception):
        await async_session.execute(
            text("UPDATE fdd_findings SET title = :title WHERE id = :id"),
            {"title": "mutated", "id": finding.id},
        )


@pytest.mark.asyncio
async def test_fdd_sections_append_only(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    section = (
        await async_session.execute(
            select(FDDSection).where(
                FDDSection.tenant_id == test_user.tenant_id,
                FDDSection.engagement_id == engagement.id,
            )
        )
    ).scalars().first()
    assert section is not None
    await async_session.execute(text(append_only_function_sql()))
    await async_session.execute(text(drop_trigger_sql("fdd_sections")))
    await async_session.execute(text(create_trigger_sql("fdd_sections")))
    await async_session.flush()
    with pytest.raises(Exception):
        await async_session.execute(
            text("UPDATE fdd_sections SET status = :status WHERE id = :id"),
            {"status": "failed", "id": section.id},
        )


# Report (4)
@pytest.mark.asyncio
async def test_get_engagement_report_structure(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    report = await get_engagement_report(async_session, test_user.tenant_id, engagement.id)
    assert {"engagement", "sections", "findings", "executive_summary", "total_ebitda_adjustments", "net_debt"}.issubset(report.keys())


@pytest.mark.asyncio
async def test_report_total_ebitda_adjustments_decimal(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    report = await get_engagement_report(async_session, test_user.tenant_id, engagement.id)
    assert isinstance(report["total_ebitda_adjustments"], Decimal)


@pytest.mark.asyncio
async def test_export_returns_xlsx_bytes(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    payload = await export_fdd_report(async_session, test_user.tenant_id, engagement.id)
    assert isinstance(payload, bytes)
    assert len(payload) > 0


@pytest.mark.asyncio
async def test_export_has_correct_sheets(async_session: AsyncSession, test_user: IamUser) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    payload = await export_fdd_report(async_session, test_user.tenant_id, engagement.id)
    wb = load_workbook(io.BytesIO(payload))
    assert "Executive Summary" in wb.sheetnames
    assert "Quality of Earnings" in wb.sheetnames
    assert "Working Capital" in wb.sheetnames


# API (3)
@pytest.mark.asyncio
async def test_create_engagement_via_api(
    async_client: AsyncClient,
    async_session: AsyncSession,
    test_user: IamUser,
    test_access_token: str,
) -> None:
    await _fund(async_session, test_user.tenant_id)
    response = await async_client.post(
        "/api/v1/advisory/fdd/engagements",
        headers={"Authorization": f"Bearer {test_access_token}"},
        json={
            "engagement_name": "Acme FDD",
            "target_company_name": "Acme",
            "analysis_period_start": "2024-01-01",
            "analysis_period_end": "2024-12-31",
            "sections_requested": [
                "quality_of_earnings",
                "working_capital",
                "debt_liability",
                "headcount",
                "revenue_quality",
            ],
        },
    )
    assert response.status_code == 201
    assert response.json()["data"]["status"] == "draft"


@pytest.mark.asyncio
async def test_run_endpoint_triggers_task(
    async_client: AsyncClient,
    async_session: AsyncSession,
    test_user: IamUser,
    test_access_token: str,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    from financeops.modules.fdd.api import routes as fdd_routes

    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)

    class _DummyAsyncResult:
        id = "dummy-task-id"

    class _DummyTask:
        @staticmethod
        def delay(*args, **kwargs):  # noqa: ANN001, ANN002
            del args, kwargs
            return _DummyAsyncResult()

    monkeypatch.setattr(fdd_routes, "run_fdd_engagement_task", _DummyTask())
    response = await async_client.post(
        f"/api/v1/advisory/fdd/engagements/{engagement.id}/run",
        headers={"Authorization": f"Bearer {test_access_token}"},
    )
    assert response.status_code == 200
    assert response.json()["data"]["task_id"] == "dummy-task-id"
    assert response.json()["data"]["status"] == "running"


@pytest.mark.asyncio
async def test_report_endpoint_returns_structure(
    async_client: AsyncClient,
    async_session: AsyncSession,
    test_user: IamUser,
    test_access_token: str,
) -> None:
    await _fund(async_session, test_user.tenant_id)
    engagement = await _create_default_engagement(async_session, test_user)
    await run_engagement(async_session, test_user.tenant_id, engagement.id)
    response = await async_client.get(
        f"/api/v1/advisory/fdd/engagements/{engagement.id}/report",
        headers={"Authorization": f"Bearer {test_access_token}"},
    )
    assert response.status_code == 200
    payload = response.json()["data"]
    assert {"engagement", "sections", "findings", "executive_summary", "total_ebitda_adjustments", "net_debt"}.issubset(payload.keys())
