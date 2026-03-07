from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook

from financeops.services.consolidation.excel_exporter import build_consolidation_excel


def _uuid(value: str) -> UUID:
    return UUID(value)


def test_excel_export_is_deterministic_and_contains_required_sheets() -> None:
    payload = dict(
        parent_currency="USD",
        consolidated_rows=[
            {
                "consolidated_result_id": _uuid("00000000-0000-0000-0000-000000000005"),
                "account_code": "3000",
                "consolidated_amount_parent": Decimal("5.000000"),
                "fx_impact_total": Decimal("0.000000"),
            },
            {
                "consolidated_result_id": _uuid("00000000-0000-0000-0000-000000000006"),
                "account_code": "4000",
                "consolidated_amount_parent": Decimal("12.000000"),
                "fx_impact_total": Decimal("0.100000"),
            },
        ],
        entity_rows=[
            {
                "account_code": "3000",
                "entity_id": _uuid("00000000-0000-0000-0000-000000000001"),
                "local_amount_total": Decimal("5.000000"),
                "parent_amount_total": Decimal("5.000000"),
                "fx_impact_total": Decimal("0.000000"),
            },
            {
                "account_code": "4000",
                "entity_id": _uuid("00000000-0000-0000-0000-000000000002"),
                "local_amount_total": Decimal("10.000000"),
                "parent_amount_total": Decimal("12.000000"),
                "fx_impact_total": Decimal("0.100000"),
            },
        ],
        line_item_rows=[
            {
                "line_item_id": _uuid("00000000-0000-0000-0000-000000000101"),
                "snapshot_line_id": _uuid("00000000-0000-0000-0000-000000000010"),
                "account_code": "4000",
                "entity_id": _uuid("00000000-0000-0000-0000-000000000002"),
                "local_currency": "EUR",
                "local_amount": Decimal("10.000000"),
                "fx_rate_used": Decimal("1.200000"),
                "parent_amount": Decimal("12.000000"),
                "fx_delta_component": Decimal("0.100000"),
            },
            {
                "line_item_id": _uuid("00000000-0000-0000-0000-000000000102"),
                "snapshot_line_id": _uuid("00000000-0000-0000-0000-000000000009"),
                "account_code": "3000",
                "entity_id": _uuid("00000000-0000-0000-0000-000000000001"),
                "local_currency": "USD",
                "local_amount": Decimal("5.000000"),
                "fx_rate_used": Decimal("1.000000"),
                "parent_amount": Decimal("5.000000"),
                "fx_delta_component": Decimal("0.000000"),
            },
        ],
        intercompany_rows=[
            {
                "pair_id": _uuid("10000000-0000-0000-0000-000000000001"),
                "entity_from": _uuid("00000000-0000-0000-0000-000000000001"),
                "entity_to": _uuid("00000000-0000-0000-0000-000000000002"),
                "account_code": "IC-100",
                "classification": "matched",
                "actual_difference": Decimal("0.000000"),
                "unexplained_difference": Decimal("0.000000"),
                "line_item_from_id": _uuid("00000000-0000-0000-0000-000000000102"),
                "line_item_to_id": _uuid("00000000-0000-0000-0000-000000000101"),
            }
        ],
        elimination_rows=[
            {
                "elimination_id": _uuid("10000000-0000-0000-0000-000000000002"),
                "intercompany_pair_id": _uuid("10000000-0000-0000-0000-000000000001"),
                "elimination_status": "applied",
                "eliminated_amount_parent": Decimal("0.000000"),
                "fx_component_impact_parent": Decimal("0.000000"),
                "residual_difference_parent": Decimal("0.000000"),
                "rule_code": "ELIM.APPLY.MATCHED",
                "reason": "matched",
            }
        ],
        unexplained_rows=[
            {
                "pair_id": _uuid("10000000-0000-0000-0000-000000000001"),
                "entity_from": _uuid("00000000-0000-0000-0000-000000000001"),
                "entity_to": _uuid("00000000-0000-0000-0000-000000000002"),
                "account_code": "IC-999",
                "classification": "unexplained",
                "expected_difference": Decimal("0.000000"),
                "actual_difference": Decimal("2.000000"),
                "fx_explained": Decimal("0.000000"),
                "unexplained_difference": Decimal("2.000000"),
                "transaction_date_from": None,
                "transaction_date_to": None,
            }
        ],
        snapshot_rows=[
            {
                "snapshot_line_id": _uuid("00000000-0000-0000-0000-000000000009"),
                "snapshot_id": _uuid("20000000-0000-0000-0000-000000000001"),
                "entity_id": _uuid("00000000-0000-0000-0000-000000000001"),
                "account_code": "3000",
                "currency": "USD",
                "local_amount": Decimal("5.000000"),
                "ic_reference": None,
                "counterparty_entity": None,
                "transaction_date": None,
                "ic_account_class": None,
                "source_artifact_reference": "snap-A",
            },
            {
                "snapshot_line_id": _uuid("00000000-0000-0000-0000-000000000010"),
                "snapshot_id": _uuid("20000000-0000-0000-0000-000000000002"),
                "entity_id": _uuid("00000000-0000-0000-0000-000000000002"),
                "account_code": "4000",
                "currency": "EUR",
                "local_amount": Decimal("10.000000"),
                "ic_reference": None,
                "counterparty_entity": None,
                "transaction_date": None,
                "ic_account_class": None,
                "source_artifact_reference": "snap-B",
            },
        ],
    )

    workbook_bytes_a, checksum_a = build_consolidation_excel(**payload)
    workbook_bytes_b, checksum_b = build_consolidation_excel(**payload)

    assert workbook_bytes_a == workbook_bytes_b
    assert checksum_a == checksum_b

    workbook = load_workbook(BytesIO(workbook_bytes_a))
    assert workbook.sheetnames == [
        "Consolidated_PnL",
        "Entity_Breakdown",
        "Line_Items",
        "Intercompany_Pairs",
        "Eliminations",
        "Unexplained",
        "Snapshot_Lines",
    ]
    first_data_row = [cell.value for cell in workbook["Consolidated_PnL"][2]]
    assert first_data_row[1] == "3000"
