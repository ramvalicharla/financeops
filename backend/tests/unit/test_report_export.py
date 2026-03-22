from __future__ import annotations

import csv
import sys
import types
from datetime import datetime
from decimal import Decimal
from io import BytesIO, StringIO

import pytest

from financeops.modules.custom_report_builder.application.export_service import (
    ReportExportService,
)


def _install_fake_weasyprint(monkeypatch: pytest.MonkeyPatch) -> None:
    class _FakeHTML:
        def __init__(self, *, string: str, base_url: str | None = None) -> None:
            self._string = string
            self._base_url = base_url

        def write_pdf(self) -> bytes:
            return b"%PDF-1.7\n%FinanceOps\n" + self._string.encode("utf-8")

    monkeypatch.setitem(sys.modules, "weasyprint", types.SimpleNamespace(HTML=_FakeHTML))


def _sample_rows() -> list[dict[str, object]]:
    return [
        {
            "metric_key": "mis.kpi.revenue",
            "metric_value": Decimal("1234.56"),
            "entity_id": "entity-1",
        },
        {
            "metric_key": "mis.kpi.ebitda",
            "metric_value": Decimal("222.22"),
            "entity_id": "entity-1",
        },
    ]


@pytest.mark.unit
def test_t_115_export_csv_returns_bytes_and_csv_filename() -> None:
    service = ReportExportService()
    payload, filename = service.export_csv(rows=_sample_rows(), report_name="Monthly Report")
    assert isinstance(payload, bytes)
    assert filename.endswith(".csv")


@pytest.mark.unit
def test_t_116_export_csv_is_valid_and_matches_header_and_row_count() -> None:
    service = ReportExportService()
    payload, _ = service.export_csv(rows=_sample_rows(), report_name="Monthly Report")
    text_payload = payload.decode("utf-8")
    reader = csv.reader(StringIO(text_payload))
    parsed = list(reader)
    assert parsed
    assert parsed[0] == ["metric_key", "metric_value", "entity_id"]
    assert len(parsed) - 1 == len(_sample_rows())


@pytest.mark.unit
def test_t_117_export_excel_returns_valid_loadable_xlsx_bytes() -> None:
    import openpyxl

    service = ReportExportService()
    payload, filename = service.export_excel(rows=_sample_rows(), report_name="Monthly Report")
    workbook = openpyxl.load_workbook(BytesIO(payload))
    assert filename.endswith(".xlsx")
    assert "Report" in workbook.sheetnames


@pytest.mark.unit
def test_t_118_export_excel_header_row_is_bold() -> None:
    import openpyxl

    service = ReportExportService()
    payload, _ = service.export_excel(rows=_sample_rows(), report_name="Monthly Report")
    workbook = openpyxl.load_workbook(BytesIO(payload))
    worksheet = workbook["Report"]
    assert worksheet["A1"].font.bold is True
    assert worksheet["B1"].font.bold is True


@pytest.mark.unit
def test_t_119_export_pdf_returns_pdf_magic_header(monkeypatch: pytest.MonkeyPatch) -> None:
    _install_fake_weasyprint(monkeypatch)
    service = ReportExportService()
    payload, filename = service.export_pdf(
        rows=_sample_rows(),
        report_name="Monthly Report",
        generated_at=datetime(2026, 1, 1, 0, 0, 0),
    )
    assert filename.endswith(".pdf")
    assert payload.startswith(b"%PDF")


@pytest.mark.unit
def test_t_120_export_csv_renders_decimal_as_string_not_float() -> None:
    service = ReportExportService()
    payload, _ = service.export_csv(rows=_sample_rows(), report_name="Monthly Report")
    text_payload = payload.decode("utf-8")
    # Decimal should remain string-rendered from Decimal, not binary float artifacts.
    assert "1,234.56" in text_payload
    assert "222.22" in text_payload


@pytest.mark.unit
def test_t_121_export_csv_empty_rows_produces_no_data_rows() -> None:
    service = ReportExportService()
    payload, filename = service.export_csv(rows=[], report_name="Monthly Report")
    assert filename.endswith(".csv")
    assert payload.decode("utf-8") == ""
