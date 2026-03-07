from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from uuid import UUID

from openpyxl import load_workbook

from financeops.services.consolidation.excel_exporter import build_consolidation_excel


def _uuid(value: str) -> UUID:
    return UUID(value)


def test_excel_export_contains_deterministic_cross_sheet_hyperlinks() -> None:
    workbook_bytes, checksum = build_consolidation_excel(
        parent_currency="USD",
        consolidated_rows=[
            {
                "consolidated_result_id": _uuid("11111111-1111-1111-1111-111111111111"),
                "account_code": "4000",
                "consolidated_amount_parent": Decimal("10.000000"),
                "fx_impact_total": Decimal("0.100000"),
            }
        ],
        entity_rows=[
            {
                "account_code": "4000",
                "entity_id": _uuid("22222222-2222-2222-2222-222222222222"),
                "local_amount_total": Decimal("10.000000"),
                "parent_amount_total": Decimal("10.000000"),
                "fx_impact_total": Decimal("0.100000"),
            }
        ],
        line_item_rows=[
            {
                "line_item_id": _uuid("33333333-3333-3333-3333-333333333333"),
                "entity_id": _uuid("22222222-2222-2222-2222-222222222222"),
                "account_code": "4000",
                "local_currency": "USD",
                "local_amount": Decimal("10.000000"),
                "fx_rate_used": Decimal("1.000000"),
                "parent_amount": Decimal("10.000000"),
                "fx_delta_component": Decimal("0.100000"),
                "snapshot_line_id": _uuid("44444444-4444-4444-4444-444444444444"),
            }
        ],
        intercompany_rows=[
            {
                "pair_id": _uuid("55555555-5555-5555-5555-555555555555"),
                "entity_from": _uuid("22222222-2222-2222-2222-222222222222"),
                "entity_to": _uuid("66666666-6666-6666-6666-666666666666"),
                "account_code": "4000",
                "classification": "matched",
                "actual_difference": Decimal("0.000000"),
                "unexplained_difference": Decimal("0.000000"),
                "line_item_from_id": _uuid("33333333-3333-3333-3333-333333333333"),
                "line_item_to_id": None,
            }
        ],
        elimination_rows=[
            {
                "elimination_id": _uuid("77777777-7777-7777-7777-777777777777"),
                "intercompany_pair_id": _uuid("55555555-5555-5555-5555-555555555555"),
                "elimination_status": "applied",
                "eliminated_amount_parent": Decimal("0.000000"),
                "fx_component_impact_parent": Decimal("0.000000"),
                "residual_difference_parent": Decimal("0.000000"),
                "rule_code": "ELIM.APPLY.MATCHED",
                "reason": "seed",
            }
        ],
        unexplained_rows=[
            {
                "pair_id": _uuid("55555555-5555-5555-5555-555555555555"),
                "entity_from": _uuid("22222222-2222-2222-2222-222222222222"),
                "entity_to": _uuid("66666666-6666-6666-6666-666666666666"),
                "account_code": "4000",
                "classification": "unexplained",
                "expected_difference": Decimal("0.000000"),
                "actual_difference": Decimal("1.000000"),
                "fx_explained": Decimal("0.000000"),
                "unexplained_difference": Decimal("1.000000"),
                "transaction_date_from": None,
                "transaction_date_to": None,
            }
        ],
        snapshot_rows=[
            {
                "snapshot_line_id": _uuid("44444444-4444-4444-4444-444444444444"),
                "snapshot_id": _uuid("88888888-8888-8888-8888-888888888888"),
                "entity_id": _uuid("22222222-2222-2222-2222-222222222222"),
                "account_code": "4000",
                "currency": "USD",
                "local_amount": Decimal("10.000000"),
                "ic_reference": "IC-1",
                "counterparty_entity": None,
                "transaction_date": None,
                "ic_account_class": "IC_RECEIVABLE",
                "source_artifact_reference": "seed",
            }
        ],
    )
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert checksum
    assert workbook.sheetnames == [
        "Consolidated_PnL",
        "Entity_Breakdown",
        "Line_Items",
        "Intercompany_Pairs",
        "Eliminations",
        "Unexplained",
        "Snapshot_Lines",
    ]
    assert str(workbook["Consolidated_PnL"]["F2"].value).startswith("=HYPERLINK(")
    assert str(workbook["Entity_Breakdown"]["F2"].value).startswith("=HYPERLINK(")
    assert str(workbook["Line_Items"]["K2"].value).startswith("=HYPERLINK(")
    assert str(workbook["Intercompany_Pairs"]["H2"].value).startswith("=HYPERLINK(")
    assert str(workbook["Eliminations"]["I2"].value).startswith("=HYPERLINK(")
