from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable
from zipfile import ZIP_STORED, ZipFile, ZipInfo

from openpyxl import Workbook
from openpyxl.styles import Font

from financeops.services.consolidation.fx_impact_calculator import quantize_output_amount
from financeops.utils.determinism import sha256_hex_bytes

_FIXED_ZIP_TIMESTAMP = (1980, 1, 1, 0, 0, 0)
_FIXED_WORKBOOK_TIMESTAMP = datetime(1980, 1, 1)

_SHEET_CONSOLIDATED = "Consolidated_PnL"
_SHEET_ENTITY = "Entity_Breakdown"
_SHEET_LINE = "Line_Items"
_SHEET_IC = "Intercompany_Pairs"
_SHEET_ELIM = "Eliminations"
_SHEET_UNEXPLAINED = "Unexplained"
_SHEET_SNAPSHOT = "Snapshot_Lines"


def _normalize_xlsx_zip(raw_bytes: bytes) -> bytes:
    source = BytesIO(raw_bytes)
    target = BytesIO()
    with ZipFile(source, "r") as src_zip:
        entries = [(name, src_zip.read(name)) for name in src_zip.namelist()]
    # Store entries uncompressed to keep byte output stable across zlib/runtime variations.
    with ZipFile(target, "w", compression=ZIP_STORED) as dst_zip:
        for name, payload in sorted(entries, key=lambda item: item[0]):
            info = ZipInfo(filename=name, date_time=_FIXED_ZIP_TIMESTAMP)
            info.compress_type = ZIP_STORED
            info.external_attr = 0o600 << 16
            dst_zip.writestr(info, payload)
    return target.getvalue()


def _sheet_header(sheet, headers: list[str]) -> None:  # type: ignore[no-untyped-def]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _as_2dp_string(value: Any) -> str:
    return str(quantize_output_amount(value))


def _hyperlink(sheet_name: str, row_index: int, label: str = "Open") -> str:
    return f'=HYPERLINK("#{sheet_name}!A{row_index}","{label}")'


def build_consolidation_excel(
    *,
    parent_currency: str,
    consolidated_rows: Iterable[dict[str, Any]],
    entity_rows: Iterable[dict[str, Any]],
    line_item_rows: Iterable[dict[str, Any]],
    intercompany_rows: Iterable[dict[str, Any]],
    elimination_rows: Iterable[dict[str, Any]],
    unexplained_rows: Iterable[dict[str, Any]],
    snapshot_rows: Iterable[dict[str, Any]],
) -> tuple[bytes, str]:
    consolidated_data = sorted(
        consolidated_rows,
        key=lambda item: (str(item["account_code"]), str(item["consolidated_result_id"])),
    )
    entity_data = sorted(
        entity_rows,
        key=lambda item: (str(item["account_code"]), str(item["entity_id"])),
    )
    line_data = sorted(
        line_item_rows,
        key=lambda item: (str(item["entity_id"]), str(item["account_code"]), str(item["line_item_id"])),
    )
    ic_data = sorted(intercompany_rows, key=lambda item: str(item["pair_id"]))
    elimination_data = sorted(elimination_rows, key=lambda item: str(item["elimination_id"]))
    unexplained_data = sorted(unexplained_rows, key=lambda item: str(item["pair_id"]))
    snapshot_data = sorted(
        snapshot_rows,
        key=lambda item: (str(item["snapshot_id"]), str(item["snapshot_line_id"])),
    )

    snapshot_row_by_id = {
        str(row["snapshot_line_id"]): idx
        for idx, row in enumerate(snapshot_data, start=2)
    }
    line_row_by_id = {
        str(row["line_item_id"]): idx
        for idx, row in enumerate(line_data, start=2)
    }
    line_first_row_by_entity_account: dict[tuple[str, str], int] = {}
    for idx, row in enumerate(line_data, start=2):
        key = (str(row["entity_id"]), str(row["account_code"]))
        line_first_row_by_entity_account.setdefault(key, idx)
    entity_first_row_by_account: dict[str, int] = {}
    for idx, row in enumerate(entity_data, start=2):
        entity_first_row_by_account.setdefault(str(row["account_code"]), idx)
    ic_row_by_pair_id = {str(row["pair_id"]): idx for idx, row in enumerate(ic_data, start=2)}

    workbook = Workbook()
    workbook.properties.created = _FIXED_WORKBOOK_TIMESTAMP
    workbook.properties.modified = _FIXED_WORKBOOK_TIMESTAMP

    sheet_consolidated = workbook.active
    sheet_consolidated.title = _SHEET_CONSOLIDATED
    _sheet_header(
        sheet_consolidated,
        [
            "Consolidated Result ID",
            "Account",
            "Parent Currency",
            "Consolidated Amount",
            "FX Impact",
            "Drill_Entity",
        ],
    )
    for row in consolidated_data:
        entity_row = entity_first_row_by_account.get(str(row["account_code"]))
        drill_cell = (
            _hyperlink(_SHEET_ENTITY, entity_row, "Entity")
            if entity_row is not None
            else ""
        )
        sheet_consolidated.append(
            [
                str(row["consolidated_result_id"]),
                str(row["account_code"]),
                parent_currency,
                _as_2dp_string(row["consolidated_amount_parent"]),
                _as_2dp_string(row["fx_impact_total"]),
                drill_cell,
            ]
        )

    sheet_entity = workbook.create_sheet(_SHEET_ENTITY)
    _sheet_header(
        sheet_entity,
        [
            "Account",
            "Entity",
            "Local Amount",
            "Parent Amount",
            "FX Impact",
            "Drill_Line_Items",
        ],
    )
    for row in entity_data:
        line_row = line_first_row_by_entity_account.get(
            (str(row["entity_id"]), str(row["account_code"]))
        )
        drill_cell = _hyperlink(_SHEET_LINE, line_row, "Line Items") if line_row is not None else ""
        sheet_entity.append(
            [
                str(row["account_code"]),
                str(row["entity_id"]),
                _as_2dp_string(row["local_amount_total"]),
                _as_2dp_string(row["parent_amount_total"]),
                _as_2dp_string(row["fx_impact_total"]),
                drill_cell,
            ]
        )

    sheet_line = workbook.create_sheet(_SHEET_LINE)
    _sheet_header(
        sheet_line,
        [
            "Line Item ID",
            "Entity",
            "Account",
            "Local Currency",
            "Local Amount",
            "FX Rate",
            "Parent Currency",
            "Parent Amount",
            "FX Impact",
            "Snapshot Line ID",
            "Drill_Snapshot",
        ],
    )
    for row in line_data:
        snapshot_row = snapshot_row_by_id.get(str(row["snapshot_line_id"]))
        drill_cell = _hyperlink(_SHEET_SNAPSHOT, snapshot_row, "Snapshot") if snapshot_row is not None else ""
        sheet_line.append(
            [
                str(row["line_item_id"]),
                str(row["entity_id"]),
                str(row["account_code"]),
                str(row["local_currency"]),
                _as_2dp_string(row["local_amount"]),
                str(row["fx_rate_used"]),
                parent_currency,
                _as_2dp_string(row["parent_amount"]),
                _as_2dp_string(row["fx_delta_component"]),
                str(row["snapshot_line_id"]),
                drill_cell,
            ]
        )

    sheet_ic = workbook.create_sheet(_SHEET_IC)
    _sheet_header(
        sheet_ic,
        [
            "Pair ID",
            "Entity From",
            "Entity To",
            "Account Code",
            "Classification",
            "Actual Difference",
            "Unexplained Difference",
            "Drill_Line_Item",
        ],
    )
    for row in ic_data:
        line_from = line_row_by_id.get(str(row.get("line_item_from_id", "")))
        line_to = line_row_by_id.get(str(row.get("line_item_to_id", "")))
        target_row = line_from or line_to
        drill_cell = _hyperlink(_SHEET_LINE, target_row, "Line Item") if target_row is not None else ""
        sheet_ic.append(
            [
                str(row["pair_id"]),
                str(row["entity_from"]),
                str(row["entity_to"]),
                str(row["account_code"]),
                str(row["classification"]),
                _as_2dp_string(row["actual_difference"]),
                _as_2dp_string(row["unexplained_difference"]),
                drill_cell,
            ]
        )

    sheet_elim = workbook.create_sheet(_SHEET_ELIM)
    _sheet_header(
        sheet_elim,
        [
            "Elimination ID",
            "Pair ID",
            "Status",
            "Eliminated Amount",
            "FX Component Impact",
            "Residual Difference",
            "Rule",
            "Reason",
            "Drill_IC_Pair",
        ],
    )
    for row in elimination_data:
        ic_row = ic_row_by_pair_id.get(str(row["intercompany_pair_id"]))
        drill_cell = _hyperlink(_SHEET_IC, ic_row, "IC Pair") if ic_row is not None else ""
        sheet_elim.append(
            [
                str(row["elimination_id"]),
                str(row["intercompany_pair_id"]),
                str(row["elimination_status"]),
                _as_2dp_string(row["eliminated_amount_parent"]),
                _as_2dp_string(row["fx_component_impact_parent"]),
                _as_2dp_string(row["residual_difference_parent"]),
                str(row["rule_code"]),
                str(row["reason"]),
                drill_cell,
            ]
        )

    sheet_unexplained = workbook.create_sheet(_SHEET_UNEXPLAINED)
    _sheet_header(
        sheet_unexplained,
        [
            "Pair ID",
            "Entity From",
            "Entity To",
            "Account Code",
            "Classification",
            "Expected Difference",
            "Actual Difference",
            "FX Explained",
            "Unexplained Difference",
            "Date From",
            "Date To",
            "Drill_IC_Pair",
        ],
    )
    for row in unexplained_data:
        ic_row = ic_row_by_pair_id.get(str(row["pair_id"]))
        drill_cell = _hyperlink(_SHEET_IC, ic_row, "IC Pair") if ic_row is not None else ""
        sheet_unexplained.append(
            [
                str(row["pair_id"]),
                str(row["entity_from"]),
                str(row["entity_to"]),
                str(row["account_code"]),
                str(row["classification"]),
                _as_2dp_string(row["expected_difference"]),
                _as_2dp_string(row["actual_difference"]),
                _as_2dp_string(row["fx_explained"]),
                _as_2dp_string(row["unexplained_difference"]),
                row["transaction_date_from"].isoformat() if row["transaction_date_from"] else "",
                row["transaction_date_to"].isoformat() if row["transaction_date_to"] else "",
                drill_cell,
            ]
        )

    sheet_snapshot = workbook.create_sheet(_SHEET_SNAPSHOT)
    _sheet_header(
        sheet_snapshot,
        [
            "Snapshot Line ID",
            "Snapshot ID",
            "Entity",
            "Account Code",
            "Currency",
            "Local Amount",
            "IC Reference",
            "Counterparty Entity",
            "Transaction Date",
            "IC Account Class",
            "Source Artifact Reference",
        ],
    )
    for row in snapshot_data:
        sheet_snapshot.append(
            [
                str(row["snapshot_line_id"]),
                str(row["snapshot_id"]),
                str(row["entity_id"]),
                str(row["account_code"]),
                str(row["currency"]),
                _as_2dp_string(row["local_amount"]),
                str(row["ic_reference"] or ""),
                str(row["counterparty_entity"] or ""),
                row["transaction_date"].isoformat() if row["transaction_date"] else "",
                str(row["ic_account_class"] or ""),
                str(row["source_artifact_reference"]),
            ]
        )

    raw = BytesIO()
    workbook.save(raw)
    normalized = _normalize_xlsx_zip(raw.getvalue())
    checksum = sha256_hex_bytes(normalized)
    return normalized, checksum

