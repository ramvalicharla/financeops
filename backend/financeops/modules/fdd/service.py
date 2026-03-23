from __future__ import annotations

import uuid
from datetime import UTC, date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from sqlalchemy import case, select
from sqlalchemy.ext.asyncio import AsyncSession

from financeops.core.exceptions import InsufficientCreditsError, NotFoundError, ValidationError
from financeops.db.models.credits import CreditReservation, ReservationStatus
from financeops.modules.fdd.models import FDDEngagement, FDDFinding, FDDSection
from financeops.modules.fdd.sections.debt_liability_review import compute_debt_review
from financeops.modules.fdd.sections.headcount_normalisation import compute_headcount_normalisation
from financeops.modules.fdd.sections.quality_of_earnings import compute_qoe
from financeops.modules.fdd.sections.revenue_quality import compute_revenue_quality
from financeops.modules.fdd.sections.working_capital_analysis import compute_wc_analysis
from financeops.services.credit_service import confirm_credits, release_credits, reserve_credits

SECTION_NAMES = {
    "quality_of_earnings",
    "working_capital",
    "debt_liability",
    "headcount",
    "revenue_quality",
}

SectionComputer = Callable[[AsyncSession, FDDEngagement], Any]
SECTION_COMPUTERS: dict[str, SectionComputer] = {
    "quality_of_earnings": compute_qoe,
    "working_capital": compute_wc_analysis,
    "debt_liability": compute_debt_review,
    "headcount": compute_headcount_normalisation,
    "revenue_quality": compute_revenue_quality,
}


def _q2(value: Decimal) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _to_decimal(value: Any, default: str = "0") -> Decimal:
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(default)


def _decimal_to_jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, list):
        return [_decimal_to_jsonable(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _decimal_to_jsonable(item) for key, item in value.items()}
    return value


async def _load_engagement(
    session: AsyncSession,
    *,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> FDDEngagement:
    engagement = (
        await session.execute(
            select(FDDEngagement).where(
                FDDEngagement.id == engagement_id,
                FDDEngagement.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if engagement is None:
        raise NotFoundError("FDD engagement not found")
    return engagement


async def _pending_reservation_id(session: AsyncSession, *, tenant_id: uuid.UUID, engagement_id: uuid.UUID) -> uuid.UUID | None:
    row = (
        await session.execute(
            select(CreditReservation.id).where(
                CreditReservation.tenant_id == tenant_id,
                CreditReservation.task_type == f"fdd:{engagement_id}",
                CreditReservation.status == ReservationStatus.pending,
            )
        )
    ).scalar_one_or_none()
    return row


def _findings_from_result(section_name: str, result: dict) -> list[dict]:
    findings = result.get("findings") if isinstance(result, dict) else None
    if isinstance(findings, list) and findings:
        return findings
    return [
        {
            "finding_type": "information",
            "severity": "informational",
            "title": f"{section_name.replace('_', ' ').title()} completed",
            "description": "Section completed with no explicit findings generated.",
            "financial_impact": None,
            "recommended_action": None,
        }
    ]


def _build_section_narrative(section_name: str, result: dict) -> str:
    title = section_name.replace("_", " ").title()
    highlights: list[str] = []
    if section_name == "quality_of_earnings":
        highlights.append(f"LTM adjusted EBITDA: {result.get('ltm_adjusted_ebitda', '0')}")
        highlights.append(f"Revenue quality score: {result.get('revenue_quality_score', '0')}")
    elif section_name == "working_capital":
        highlights.append(f"NWC peg: {result.get('nwc_peg', '0')}")
    elif section_name == "debt_liability":
        highlights.append(f"Net debt: {result.get('net_debt', '0')}")
    elif section_name == "headcount":
        highlights.append(f"Normalisation adjustment: {result.get('normalisation_adjustment', '0')}")
    elif section_name == "revenue_quality":
        highlights.append(f"Quality score: {result.get('quality_score', '0')}")
    body = "; ".join(str(item) for item in highlights if item)
    return f"{title} analysis complete. {body}".strip()


async def create_engagement(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_name: str,
    target_company_name: str,
    analysis_period_start: date,
    analysis_period_end: date,
    sections_requested: list[str],
    created_by: uuid.UUID,
) -> FDDEngagement:
    """
    Create FDD engagement and reserve credits.
    """
    if analysis_period_end < analysis_period_start:
        raise ValidationError("analysis_period_end must be on or after analysis_period_start")
    cleaned_sections = [section for section in sections_requested if section in SECTION_NAMES]
    if not cleaned_sections:
        raise ValidationError("sections_requested must include at least one valid section")

    engagement = FDDEngagement(
        tenant_id=tenant_id,
        engagement_name=engagement_name,
        target_company_name=target_company_name,
        analysis_period_start=analysis_period_start,
        analysis_period_end=analysis_period_end,
        status="draft",
        credit_cost=2500,
        sections_requested=cleaned_sections,
        sections_completed=[],
        created_by=created_by,
    )
    session.add(engagement)
    await session.flush()

    reservation_id = await reserve_credits(
        session,
        tenant_id=tenant_id,
        task_type=f"fdd:{engagement.id}",
        amount=Decimal("2500"),
    )
    if reservation_id is None:
        raise InsufficientCreditsError("Unable to reserve credits")
    engagement.credits_reserved_at = datetime.now(UTC)
    await session.flush()
    return engagement


async def run_engagement(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> FDDEngagement:
    """
    Run all requested sections for this engagement.
    """
    engagement = await _load_engagement(session, tenant_id=tenant_id, engagement_id=engagement_id)
    engagement.status = "running"
    engagement.updated_at = datetime.now(UTC)
    await session.flush()

    completed_sections: list[str] = []
    failed_sections: list[str] = []

    for section_name in list(engagement.sections_requested or []):
        computer = SECTION_COMPUTERS.get(section_name)
        if computer is None:
            failed_sections.append(section_name)
            section_row = FDDSection(
                engagement_id=engagement.id,
                tenant_id=tenant_id,
                section_name=section_name,
                status="failed",
                result_data={"error": "unsupported_section"},
                ai_narrative=f"{section_name} is not supported by the current runner.",
                duration_seconds=Decimal("0.00"),
            )
            session.add(section_row)
            await session.flush()
            continue

        started_at = datetime.now(UTC)
        try:
            result_raw = await computer(session, engagement)
            result_json = _decimal_to_jsonable(result_raw)
            ai_narrative = _build_section_narrative(section_name, result_json if isinstance(result_json, dict) else {})
            section_row = FDDSection(
                engagement_id=engagement.id,
                tenant_id=tenant_id,
                section_name=section_name,
                status="completed",
                result_data=result_json if isinstance(result_json, dict) else {"result": result_json},
                ai_narrative=ai_narrative,
                duration_seconds=_q2(Decimal(str((datetime.now(UTC) - started_at).total_seconds()))),
            )
            session.add(section_row)
            await session.flush()

            for finding in _findings_from_result(section_name, result_raw if isinstance(result_raw, dict) else {}):
                session.add(
                    FDDFinding(
                        engagement_id=engagement.id,
                        section_id=section_row.id,
                        tenant_id=tenant_id,
                        finding_type=str(finding.get("finding_type") or "information"),
                        severity=str(finding.get("severity") or "informational"),
                        title=str(finding.get("title") or "FDD finding"),
                        description=str(finding.get("description") or ""),
                        financial_impact=_to_decimal(finding.get("financial_impact")) if finding.get("financial_impact") is not None else None,
                        financial_impact_currency="INR",
                        recommended_action=(
                            str(finding.get("recommended_action"))
                            if finding.get("recommended_action") is not None
                            else None
                        ),
                    )
                )
            completed_sections.append(section_name)
        except Exception as exc:  # noqa: BLE001
            failed_sections.append(section_name)
            section_row = FDDSection(
                engagement_id=engagement.id,
                tenant_id=tenant_id,
                section_name=section_name,
                status="failed",
                result_data={"error": str(exc)},
                ai_narrative=f"{section_name} failed: {exc}",
                duration_seconds=_q2(Decimal(str((datetime.now(UTC) - started_at).total_seconds()))),
            )
            session.add(section_row)
            await session.flush()

    engagement.sections_completed = completed_sections

    reservation_id = await _pending_reservation_id(
        session,
        tenant_id=tenant_id,
        engagement_id=engagement.id,
    )

    if completed_sections:
        if reservation_id is not None:
            await confirm_credits(
                session,
                tenant_id=tenant_id,
                reservation_id=reservation_id,
                user_id=engagement.created_by,
            )
        engagement.credits_deducted_at = datetime.now(UTC)
        engagement.status = "completed"
    else:
        if reservation_id is not None:
            await release_credits(
                session,
                tenant_id=tenant_id,
                reservation_id=reservation_id,
                user_id=engagement.created_by,
            )
        engagement.status = "failed"

    engagement.updated_at = datetime.now(UTC)
    await session.flush()
    return engagement


async def get_engagement_report(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> dict:
    """
    Assemble full FDD report.
    """
    engagement = await _load_engagement(session, tenant_id=tenant_id, engagement_id=engagement_id)

    section_rows = (
        await session.execute(
            select(FDDSection).where(
                FDDSection.tenant_id == tenant_id,
                FDDSection.engagement_id == engagement.id,
            )
        )
    ).scalars().all()
    latest_by_name: dict[str, FDDSection] = {}
    for row in sorted(section_rows, key=lambda item: (item.section_name, item.computed_at, item.id)):
        latest_by_name[row.section_name] = row

    severity_rank = case(
        (FDDFinding.severity == "critical", 1),
        (FDDFinding.severity == "high", 2),
        (FDDFinding.severity == "medium", 3),
        (FDDFinding.severity == "low", 4),
        else_=5,
    )
    findings = (
        await session.execute(
            select(FDDFinding)
            .where(
                FDDFinding.tenant_id == tenant_id,
                FDDFinding.engagement_id == engagement.id,
            )
            .order_by(severity_rank.asc(), FDDFinding.created_at.asc())
        )
    ).scalars().all()

    qoe_result = (latest_by_name.get("quality_of_earnings").result_data if latest_by_name.get("quality_of_earnings") else {}) or {}
    debt_result = (latest_by_name.get("debt_liability").result_data if latest_by_name.get("debt_liability") else {}) or {}
    wc_result = (latest_by_name.get("working_capital").result_data if latest_by_name.get("working_capital") else {}) or {}

    adjustments_total = Decimal("0")
    for item in qoe_result.get("adjustments", []) if isinstance(qoe_result, dict) else []:
        if isinstance(item, dict):
            adjustments_total += _to_decimal(item.get("amount"))

    net_debt = _to_decimal(debt_result.get("net_debt")) if isinstance(debt_result, dict) else Decimal("0")
    wc_adjustment = _to_decimal(wc_result.get("adjustment_recommendation")) if isinstance(wc_result, dict) else Decimal("0")
    ltm_adjusted_ebitda = _to_decimal(qoe_result.get("ltm_adjusted_ebitda")) if isinstance(qoe_result, dict) else Decimal("0")
    price_adjustments = _q2(net_debt + wc_adjustment)

    executive_summary = (
        f"Engagement {engagement.engagement_name} completed with {len(latest_by_name)} sections and "
        f"{len(findings)} findings. LTM adjusted EBITDA {ltm_adjusted_ebitda}, net debt {net_debt}."
    )

    return {
        "engagement": engagement,
        "sections": {name: row.result_data for name, row in latest_by_name.items()},
        "findings": findings,
        "executive_summary": executive_summary,
        "total_ebitda_adjustments": _q2(adjustments_total),
        "net_debt": _q2(net_debt),
        "recommended_price_adjustments": price_adjustments,
        "ltm_adjusted_ebitda": _q2(ltm_adjusted_ebitda),
    }


async def export_fdd_report(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> bytes:
    """
    Export FDD report as Excel workbook.
    """
    report = await get_engagement_report(
        session,
        tenant_id=tenant_id,
        engagement_id=engagement_id,
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Executive Summary"
    sheet.append(["Engagement", report["engagement"].engagement_name])
    sheet.append(["Target", report["engagement"].target_company_name])
    sheet.append(["Executive Summary", report["executive_summary"]])
    sheet.append(["LTM Adjusted EBITDA", format(report["ltm_adjusted_ebitda"], "f")])
    sheet.append(["Net Debt", format(report["net_debt"], "f")])
    sheet.append(["Recommended Price Adjustments", format(report["recommended_price_adjustments"], "f")])

    section_sheet_names = {
        "quality_of_earnings": "Quality of Earnings",
        "working_capital": "Working Capital",
        "debt_liability": "Debt Review",
        "headcount": "Headcount",
        "revenue_quality": "Revenue Quality",
    }
    sections = report.get("sections", {})
    for key, sheet_name in section_sheet_names.items():
        ws = workbook.create_sheet(sheet_name)
        payload = sections.get(key, {}) if isinstance(sections, dict) else {}
        ws.append(["Section", sheet_name])
        if isinstance(payload, dict):
            for item_key, item_value in payload.items():
                ws.append([str(item_key), str(item_value)])
        else:
            ws.append(["result", str(payload)])

    findings_sheet = workbook.create_sheet("Findings")
    findings_sheet.append(["Severity", "Type", "Title", "Financial Impact"])
    for finding in report.get("findings", []):
        impact = finding.financial_impact if getattr(finding, "financial_impact", None) is not None else ""
        findings_sheet.append([
            finding.severity,
            finding.finding_type,
            finding.title,
            str(impact),
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__ = [
    "create_engagement",
    "run_engagement",
    "get_engagement_report",
    "export_fdd_report",
]
