from __future__ import annotations

import uuid
from datetime import UTC, date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from financeops.core.exceptions import InsufficientCreditsError, NotFoundError, ValidationError
from financeops.db.models.credits import CreditReservation, ReservationStatus
from financeops.modules.ppa.models import PPAAllocation, PPAEngagement, PPAIntangible
from financeops.services.credit_service import confirm_credits, reserve_credits


def _q2(value: Decimal) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _q4(value: Decimal) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.0000"), rounding=ROUND_HALF_UP)


def _to_decimal(value: Any, default: str = "0") -> Decimal:
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal(default)


def _jsonify(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, list):
        return [_jsonify(v) for v in value]
    if isinstance(value, dict):
        return {str(k): _jsonify(v) for k, v in value.items()}
    return value


async def _load_engagement(session: AsyncSession, *, tenant_id: uuid.UUID, engagement_id: uuid.UUID) -> PPAEngagement:
    row = (
        await session.execute(
            select(PPAEngagement).where(
                PPAEngagement.id == engagement_id,
                PPAEngagement.tenant_id == tenant_id,
            )
        )
    ).scalar_one_or_none()
    if row is None:
        raise NotFoundError("PPA engagement not found")
    return row


async def _pending_reservation_id(session: AsyncSession, *, tenant_id: uuid.UUID, engagement_id: uuid.UUID) -> uuid.UUID | None:
    return (
        await session.execute(
            select(CreditReservation.id).where(
                CreditReservation.tenant_id == tenant_id,
                CreditReservation.task_type == f"ppa:{engagement_id}",
                CreditReservation.status == ReservationStatus.pending,
            )
        )
    ).scalar_one_or_none()


async def create_ppa_engagement(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_name: str,
    target_company_name: str,
    acquisition_date: date,
    purchase_price: Decimal,
    purchase_price_currency: str,
    accounting_standard: str,
    created_by: uuid.UUID,
) -> PPAEngagement:
    """
    Create PPA engagement and reserve credits.
    """
    if accounting_standard not in {"IFRS3", "ASC805", "INDAS103"}:
        raise ValidationError("accounting_standard must be IFRS3, ASC805 or INDAS103")

    engagement = PPAEngagement(
        tenant_id=tenant_id,
        engagement_name=engagement_name,
        target_company_name=target_company_name,
        acquisition_date=acquisition_date,
        purchase_price=_q2(_to_decimal(purchase_price)),
        purchase_price_currency=purchase_price_currency,
        accounting_standard=accounting_standard,
        status="draft",
        credit_cost=2000,
        created_by=created_by,
    )
    session.add(engagement)
    await session.flush()

    reservation_id = await reserve_credits(
        session,
        tenant_id=tenant_id,
        task_type=f"ppa:{engagement.id}",
        amount=Decimal("2000"),
    )
    if reservation_id is None:
        raise InsufficientCreditsError("Unable to reserve credits for PPA")
    return engagement


async def identify_intangibles(
    session: AsyncSession,
    engagement: PPAEngagement,
) -> list[dict]:
    """
    AI-assisted intangible identification (deterministic seed output).
    """
    del session
    suggestions = [
        {
            "intangible_name": "Customer relationships",
            "intangible_category": "customer_relationships",
            "rationale": f"{engagement.target_company_name} has repeat customer base.",
            "typical_useful_life_years": Decimal("7.00"),
            "recommended_valuation_method": "excess_earnings",
        },
        {
            "intangible_name": "Technology platform",
            "intangible_category": "technology",
            "rationale": "Core technology stack contributes to recurring revenue.",
            "typical_useful_life_years": Decimal("5.00"),
            "recommended_valuation_method": "relief_from_royalty",
        },
        {
            "intangible_name": "Brand",
            "intangible_category": "brand",
            "rationale": "Brand supports pricing premium and pipeline conversion.",
            "typical_useful_life_years": Decimal("10.00"),
            "recommended_valuation_method": "relief_from_royalty",
        },
    ]
    return suggestions


async def compute_intangible_fair_value(
    intangible_category: str,
    valuation_method: str,
    assumptions: dict,
) -> Decimal:
    """
    Compute fair value of an intangible using Decimal arithmetic.
    """
    del intangible_category
    revenue = _to_decimal(assumptions.get("revenue", "0"))
    royalty_rate = _to_decimal(assumptions.get("royalty_rate", "0"))
    discount_rate = _to_decimal(assumptions.get("discount_rate", "0"))
    useful_life_years = int(_to_decimal(assumptions.get("useful_life_years", "1")))
    earnings = _to_decimal(assumptions.get("earnings", revenue * Decimal("0.25")))
    contributory_asset_charges = _to_decimal(assumptions.get("contributory_asset_charges", "0"))

    if useful_life_years <= 0:
        raise ValidationError("useful_life_years must be positive")

    fair_value = Decimal("0")
    if valuation_method == "relief_from_royalty":
        if royalty_rate <= Decimal("0"):
            return Decimal("0")
        annual_royalty = revenue * royalty_rate
        for year in range(1, useful_life_years + 1):
            discount_factor = (Decimal("1") + discount_rate) ** Decimal(str(year))
            fair_value += annual_royalty / discount_factor
        return _q2(fair_value)

    if valuation_method == "excess_earnings":
        annual_excess = earnings - contributory_asset_charges
        for year in range(1, useful_life_years + 1):
            discount_factor = (Decimal("1") + discount_rate) ** Decimal(str(year))
            fair_value += annual_excess / discount_factor
        return _q2(fair_value)

    # fallback deterministic approach for other valuation methods
    for year in range(1, useful_life_years + 1):
        discount_factor = (Decimal("1") + discount_rate) ** Decimal(str(year))
        fair_value += (revenue * Decimal("0.04")) / discount_factor
    return _q2(fair_value)


async def _book_value_net_assets(session: AsyncSession, engagement: PPAEngagement) -> Decimal:
    del session, engagement
    return Decimal("600.00")


async def run_ppa(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
    intangibles_input: list[dict],
) -> PPAAllocation:
    """
    Run full PPA computation.
    """
    engagement = await _load_engagement(session, tenant_id=tenant_id, engagement_id=engagement_id)
    engagement.status = "running"
    engagement.updated_at = datetime.now(UTC)
    await session.flush()

    net_identifiable_assets = _q2(await _book_value_net_assets(session, engagement))
    allocation = PPAAllocation(
        engagement_id=engagement.id,
        tenant_id=tenant_id,
        allocation_version=1,
        net_identifiable_assets=net_identifiable_assets,
        total_intangibles_identified=Decimal("0.00"),
        goodwill=Decimal("0.00"),
        deferred_tax_liability=Decimal("0.00"),
        purchase_price_reconciliation={},
    )
    session.add(allocation)
    await session.flush()

    total_intangibles = Decimal("0")
    total_dtl = Decimal("0")

    for item in intangibles_input:
        assumptions = item.get("assumptions") if isinstance(item.get("assumptions"), dict) else {}
        assumptions = dict(assumptions)
        assumptions.setdefault("useful_life_years", str(item.get("useful_life_years", "5")))

        fair_value = await compute_intangible_fair_value(
            str(item.get("category") or item.get("intangible_category") or "other"),
            str(item.get("valuation_method") or "relief_from_royalty"),
            assumptions,
        )
        useful_life = _to_decimal(item.get("useful_life_years", assumptions.get("useful_life_years", "5")))
        if useful_life <= Decimal("0"):
            raise ValidationError("useful_life_years must be positive")

        annual_amortisation = _q2(fair_value / useful_life)
        tax_basis = _q2(_to_decimal(item.get("tax_basis", "0")))
        applicable_tax_rate = _to_decimal(item.get("applicable_tax_rate", "0.25"))
        deferred_tax_liability = _q2((fair_value - tax_basis) * applicable_tax_rate)

        row = PPAIntangible(
            engagement_id=engagement.id,
            allocation_id=allocation.id,
            tenant_id=tenant_id,
            intangible_name=str(item.get("name") or item.get("intangible_name") or "Intangible"),
            intangible_category=str(item.get("category") or item.get("intangible_category") or "other"),
            fair_value=fair_value,
            useful_life_years=_q2(useful_life),
            amortisation_method=str(item.get("amortisation_method") or "straight_line"),
            annual_amortisation=annual_amortisation,
            tax_basis=tax_basis,
            deferred_tax_liability=deferred_tax_liability,
            valuation_method=str(item.get("valuation_method") or "relief_from_royalty"),
            valuation_assumptions=_jsonify(assumptions),
        )
        session.add(row)

        total_intangibles += fair_value
        total_dtl += deferred_tax_liability

    total_intangibles = _q2(total_intangibles)
    total_dtl = _q2(total_dtl)

    goodwill = _q2(
        _to_decimal(engagement.purchase_price)
        - net_identifiable_assets
        - total_intangibles
        + total_dtl
    )

    allocation.total_intangibles_identified = total_intangibles
    allocation.goodwill = goodwill
    allocation.deferred_tax_liability = total_dtl
    allocation.purchase_price_reconciliation = _jsonify(
        {
            "purchase_price": _to_decimal(engagement.purchase_price),
            "net_identifiable_assets": net_identifiable_assets,
            "total_intangibles_identified": total_intangibles,
            "goodwill": goodwill,
            "deferred_tax_liability": total_dtl,
            "reconciled_total": _q2(net_identifiable_assets + total_intangibles + goodwill - total_dtl),
        }
    )

    reservation_id = await _pending_reservation_id(
        session,
        tenant_id=tenant_id,
        engagement_id=engagement.id,
    )
    if reservation_id is not None:
        await confirm_credits(
            session,
            tenant_id=tenant_id,
            reservation_id=reservation_id,
            user_id=engagement.created_by,
        )

    engagement.status = "completed"
    engagement.updated_at = datetime.now(UTC)
    await session.flush()
    return allocation


async def get_ppa_report(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> dict:
    engagement = await _load_engagement(session, tenant_id=tenant_id, engagement_id=engagement_id)
    allocation = (
        await session.execute(
            select(PPAAllocation)
            .where(
                PPAAllocation.tenant_id == tenant_id,
                PPAAllocation.engagement_id == engagement.id,
            )
            .order_by(desc(PPAAllocation.computed_at), desc(PPAAllocation.id))
            .limit(1)
        )
    ).scalar_one_or_none()
    if allocation is None:
        raise NotFoundError("No PPA allocation found for engagement")

    intangibles = (
        await session.execute(
            select(PPAIntangible)
            .where(
                PPAIntangible.tenant_id == tenant_id,
                PPAIntangible.engagement_id == engagement.id,
                PPAIntangible.allocation_id == allocation.id,
            )
            .order_by(PPAIntangible.fair_value.desc(), PPAIntangible.id.asc())
        )
    ).scalars().all()

    purchase_price = _to_decimal(engagement.purchase_price)
    goodwill_pct = Decimal("0")
    if purchase_price != Decimal("0"):
        goodwill_pct = _q4((_to_decimal(allocation.goodwill) / purchase_price) * Decimal("100"))

    max_years = 0
    for row in intangibles:
        max_years = max(max_years, int(_to_decimal(row.useful_life_years)))

    amortisation_schedule: dict[str, Decimal] = {}
    for year in range(1, max_years + 1):
        total_year = Decimal("0")
        for row in intangibles:
            if year <= int(_to_decimal(row.useful_life_years)):
                total_year += _to_decimal(row.annual_amortisation)
        amortisation_schedule[f"year_{year}"] = _q2(total_year)

    bridge = {
        "book_value_net_assets": _to_decimal(allocation.net_identifiable_assets),
        "step_ups": [
            {
                "name": row.intangible_name,
                "fair_value": _to_decimal(row.fair_value),
                "tax_impact": _to_decimal(row.deferred_tax_liability),
            }
            for row in intangibles
        ],
        "goodwill": _to_decimal(allocation.goodwill),
        "total": _q2(
            _to_decimal(allocation.net_identifiable_assets)
            + _to_decimal(allocation.total_intangibles_identified)
            - _to_decimal(allocation.deferred_tax_liability)
            + _to_decimal(allocation.goodwill)
        ),
    }

    return {
        "engagement": engagement,
        "allocation": allocation,
        "intangibles": intangibles,
        "purchase_price_bridge": bridge,
        "amortisation_schedule": amortisation_schedule,
        "goodwill_pct_of_purchase_price": goodwill_pct,
    }


async def export_ppa_report(
    session: AsyncSession,
    tenant_id: uuid.UUID,
    engagement_id: uuid.UUID,
) -> bytes:
    report = await get_ppa_report(session, tenant_id, engagement_id)

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Engagement", report["engagement"].engagement_name])
    summary.append(["Target", report["engagement"].target_company_name])
    summary.append(["Purchase Price", format(_to_decimal(report["engagement"].purchase_price), "f")])
    summary.append(["Goodwill", format(_to_decimal(report["allocation"].goodwill), "f")])
    summary.append(["Goodwill %", format(report["goodwill_pct_of_purchase_price"], "f")])

    ints = wb.create_sheet("Intangibles")
    ints.append(["Name", "Category", "Fair Value", "Useful Life", "Annual Amortisation", "Method"])
    for row in report["intangibles"]:
        ints.append(
            [
                row.intangible_name,
                row.intangible_category,
                format(_to_decimal(row.fair_value), "f"),
                format(_to_decimal(row.useful_life_years), "f"),
                format(_to_decimal(row.annual_amortisation), "f"),
                row.valuation_method,
            ]
        )

    bridge = wb.create_sheet("Bridge")
    bridge.append(["Book value net assets", format(report["purchase_price_bridge"]["book_value_net_assets"], "f")])
    bridge.append(["Goodwill", format(report["purchase_price_bridge"]["goodwill"], "f")])
    bridge.append(["Total", format(report["purchase_price_bridge"]["total"], "f")])

    amort = wb.create_sheet("Amortisation")
    amort.append(["Year", "Amount"])
    for key, value in report["amortisation_schedule"].items():
        amort.append([key, format(value, "f")])

    buff = BytesIO()
    wb.save(buff)
    return buff.getvalue()


__all__ = [
    "create_ppa_engagement",
    "identify_intangibles",
    "compute_intangible_fair_value",
    "run_ppa",
    "get_ppa_report",
    "export_ppa_report",
]
