from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


def parse_excel_bytes(
    content: bytes, *, preferred_sheet: str | None = None
) -> dict[str, Any]:
    workbook = load_workbook(filename=BytesIO(content), data_only=False, read_only=True)
    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    if not visible_sheets:
        raise ValueError("No visible sheets found in workbook")

    sheet = _pick_sheet(visible_sheets, preferred_sheet)
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if value is None else str(value).strip() for value in row])
    non_empty_rows = [row for row in rows if any(str(cell).strip() for cell in row)]
    headers = [str(item).strip() for item in (non_empty_rows[0] if non_empty_rows else [])]
    data_rows = non_empty_rows[1:] if len(non_empty_rows) > 1 else []

    values = [str(cell).strip() for row in data_rows for cell in row if str(cell).strip()]
    numeric_count = sum(1 for value in values if _is_number(value))
    text_count = sum(1 for value in values if not _is_number(value))
    formula_count = sum(1 for value in values if value.startswith("="))
    total_cells = max(len(values), 1)
    row_labels = [str(row[0]).strip() for row in data_rows if row and str(row[0]).strip()]

    return {
        "sheet_name": sheet.title,
        "headers": headers,
        "rows": data_rows,
        "row_labels": row_labels,
        "header_row_index": 0,
        "data_start_row_index": 1 if headers else 0,
        "blank_row_density": Decimal(len(rows) - len(non_empty_rows)) / Decimal(max(len(rows), 1)),
        "formula_density": Decimal(formula_count) / Decimal(total_cells),
        "text_to_numeric_ratio": Decimal(text_count) / Decimal(max(numeric_count, 1)),
        "merged_cell_count": len(sheet.merged_cells.ranges),
    }


def _pick_sheet(visible_sheets: list, preferred_sheet: str | None):
    if preferred_sheet:
        for sheet in visible_sheets:
            if sheet.title == preferred_sheet:
                return sheet
    return visible_sheets[0]


def _is_number(value: str) -> bool:
    try:
        Decimal(value.replace(",", ""))
        return True
    except (TypeError, ValueError, InvalidOperation):
        return False


