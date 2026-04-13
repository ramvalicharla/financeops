from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import re
import uuid
from collections import defaultdict
from datetime import UTC, datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import func, or_, select, update
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from financeops.core.exceptions import AuthorizationError, NotFoundError, ValidationError
from financeops.core.governance.airlock import AirlockAdmissionService
from financeops.modules.coa.models import (
    CoaAccountGroup,
    CoaAccountSubgroup,
    CoaIndustryTemplate,
    CoaLedgerAccount,
    CoaSourceType,
    TenantCoaAccount,
    CoaUploadBatch,
    CoaUploadMode,
    CoaUploadStagingRow,
    CoaUploadStatus,
)
from financeops.modules.coa.application.tenant_coa_service import TenantCoaService

_REQUIRED_COLUMNS = (
    "group_code",
    "group_name",
    "subgroup_code",
    "subgroup_name",
    "ledger_code",
    "ledger_name",
    "ledger_type",
    "is_control_account",
)
_ALLOWED_LEDGER_TYPES = {"ASSET", "LIABILITY", "INCOME", "EXPENSE"}
_MAX_UPLOAD_BYTES = 5 * 1024 * 1024
_ACCOUNT_ALIASES = ("account", "particulars", "ledger", "name")
_DEBIT_ALIASES = ("debit", "dr")
_CREDIT_ALIASES = ("credit", "cr")
_FLEXIBLE_UPLOAD_KIND = "FLEXIBLE_TB"
log = logging.getLogger(__name__)


class CoaUploadService:
    def __init__(self, session: AsyncSession) -> None:
        self._session = session

    @staticmethod
    def _normalise_header(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _read_dataframe(*, file_name: str, file_bytes: bytes) -> pd.DataFrame:
        lower_name = file_name.lower()
        if lower_name.endswith(".csv"):
            return pd.read_csv(io.BytesIO(file_bytes), dtype=str, keep_default_na=False)
        if lower_name.endswith(".xlsx"):
            return pd.read_excel(io.BytesIO(file_bytes), dtype=str)
        raise ValidationError("Only .csv or .xlsx files are supported")

    @classmethod
    def _match_alias(cls, columns: list[str], aliases: tuple[str, ...]) -> str | None:
        alias_set = {alias.lower() for alias in aliases}
        for column in columns:
            if cls._normalise_header(column) in alias_set:
                return column
        return None

    @staticmethod
    def _clean_numeric_series(series: pd.Series) -> pd.Series:
        cleaned = (
            series.fillna("")
            .astype(str)
            .str.strip()
            .str.replace(",", "", regex=False)
            .str.replace(r"^\((.*)\)$", r"-\1", regex=True)
        )
        numeric = pd.to_numeric(cleaned.replace("", pd.NA), errors="coerce")
        return numeric

    def parse_normalized_dataframe(self, *, file_name: str, file_bytes: bytes) -> pd.DataFrame:
        if len(file_bytes) > _MAX_UPLOAD_BYTES:
            raise ValidationError("File size exceeds 5MB limit")

        dataframe = self._read_dataframe(file_name=file_name, file_bytes=file_bytes).copy()
        dataframe.columns = [str(column or "").strip() for column in dataframe.columns]
        original_columns = list(dataframe.columns)
        normalized_columns = [self._normalise_header(column) for column in original_columns]
        if not any(normalized_columns):
            raise ValidationError("Uploaded file has no usable columns")

        if set(_REQUIRED_COLUMNS).issubset(set(normalized_columns)):
            dataframe.columns = normalized_columns
            for column in dataframe.columns:
                if dataframe[column].dtype == object:
                    dataframe[column] = dataframe[column].fillna("").astype(str).str.strip()
            dataframe = dataframe.replace("", pd.NA).dropna(how="all").fillna("")
            return dataframe.reset_index(drop=True)

        account_column = self._match_alias(original_columns, _ACCOUNT_ALIASES)
        debit_column = self._match_alias(original_columns, _DEBIT_ALIASES)
        credit_column = self._match_alias(original_columns, _CREDIT_ALIASES)

        if account_column is None:
            raise ValidationError("account column is required")
        if debit_column is None and credit_column is None:
            raise ValidationError("at least one of debit or credit columns is required")

        normalized = pd.DataFrame()
        normalized["account"] = (
            dataframe[account_column].fillna("").astype(str).str.strip()
        )
        if debit_column is not None:
            normalized["debit"] = self._clean_numeric_series(dataframe[debit_column])
        if credit_column is not None:
            normalized["credit"] = self._clean_numeric_series(dataframe[credit_column])

        if "debit" not in normalized.columns:
            normalized["debit"] = 0.0
        if "credit" not in normalized.columns:
            normalized["credit"] = 0.0

        normalized = normalized.loc[
            ~(
                normalized["account"].eq("")
                & normalized["debit"].fillna(0).eq(0)
                & normalized["credit"].fillna(0).eq(0)
            )
        ]
        normalized["debit"] = normalized["debit"].fillna(0)
        normalized["credit"] = normalized["credit"].fillna(0)
        return normalized.reset_index(drop=True)

    @staticmethod
    def _normalise_bool(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        text = str(value or "").strip().lower()
        return text in {"1", "true", "yes", "y"}

    @staticmethod
    def _normalise_trial_balance_amount(value: Any) -> Decimal:
        return Decimal(str(value or "0")).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)

    @staticmethod
    def _build_flexible_plan_item(
        *,
        row_number: int,
        row: dict[str, Any],
        status: str,
        **extra: Any,
    ) -> dict[str, Any]:
        item = {
            "row_number": row_number,
            "account": str(row.get("account") or "").strip(),
            "debit": CoaUploadService._normalise_trial_balance_amount(row.get("debit")),
            "credit": CoaUploadService._normalise_trial_balance_amount(row.get("credit")),
            "status": status,
        }
        item.update(extra)
        return item

    @staticmethod
    def _json_safe_value(value: Any) -> Any:
        if isinstance(value, Decimal):
            return format(value, "f")
        if isinstance(value, dict):
            return {str(key): CoaUploadService._json_safe_value(item) for key, item in value.items()}
        if isinstance(value, list):
            return [CoaUploadService._json_safe_value(item) for item in value]
        return value

    @staticmethod
    def _normalise_row(raw: dict[str, Any], row_number: int) -> dict[str, Any]:
        return {
            "row_number": row_number,
            "group_code": str(raw.get("group_code", "")).strip().upper(),
            "group_name": str(raw.get("group_name", "")).strip(),
            "subgroup_code": str(raw.get("subgroup_code", "")).strip().upper(),
            "subgroup_name": str(raw.get("subgroup_name", "")).strip(),
            "ledger_code": str(raw.get("ledger_code", "")).strip().upper(),
            "ledger_name": str(raw.get("ledger_name", "")).strip(),
            "ledger_type": str(raw.get("ledger_type", "")).strip().upper(),
            "is_control_account": CoaUploadService._normalise_bool(raw.get("is_control_account", False)),
        }

    def parse_file(self, *, file_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
        if len(file_bytes) > _MAX_UPLOAD_BYTES:
            raise ValidationError("File size exceeds 5MB limit")

        lower_name = file_name.lower()
        rows: list[dict[str, Any]] = []

        if lower_name.endswith(".csv"):
            decoded = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            for row_number, row in enumerate(reader, start=2):
                rows.append(self._normalise_row(row, row_number))
            return rows

        if lower_name.endswith(".xlsx"):
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            worksheet = workbook.active
            values = worksheet.iter_rows(values_only=True)
            headers_row = next(values, None)
            if headers_row is None:
                return []
            headers = [str(item or "").strip() for item in headers_row]
            for row_number, row_values in enumerate(values, start=2):
                raw_row = {headers[index]: row_values[index] for index in range(len(headers))}
                rows.append(self._normalise_row(raw_row, row_number))
            return rows

        raise ValidationError("Only .csv or .xlsx files are supported")

    def validate_structure(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        errors: list[dict[str, Any]] = []
        group_name_by_code: dict[str, str] = {}
        subgroup_group_by_code: dict[str, str] = {}
        ledger_subgroup_by_code: dict[str, str] = {}
        seen_ledger_codes: set[str] = set()

        for row in rows:
            row_errors: list[str] = []
            row_number = int(row["row_number"])

            for column in _REQUIRED_COLUMNS:
                value = row.get(column)
                if value is None or (isinstance(value, str) and not value.strip()):
                    row_errors.append(f"{column} is required")

            ledger_type = str(row.get("ledger_type", "")).upper()
            if ledger_type and ledger_type not in _ALLOWED_LEDGER_TYPES:
                row_errors.append("ledger_type must be one of ASSET, LIABILITY, INCOME, EXPENSE")

            group_code = str(row.get("group_code", ""))
            group_name = str(row.get("group_name", ""))
            if group_code:
                previous_group_name = group_name_by_code.get(group_code)
                if previous_group_name and previous_group_name != group_name:
                    row_errors.append("group_code maps to multiple group_name values")
                group_name_by_code[group_code] = group_name

            subgroup_code = str(row.get("subgroup_code", ""))
            if subgroup_code:
                previous_group_code = subgroup_group_by_code.get(subgroup_code)
                if previous_group_code and previous_group_code != group_code:
                    row_errors.append("subgroup_code must belong to only one group_code")
                subgroup_group_by_code[subgroup_code] = group_code

            ledger_code = str(row.get("ledger_code", ""))
            if ledger_code:
                previous_subgroup = ledger_subgroup_by_code.get(ledger_code)
                if previous_subgroup and previous_subgroup != subgroup_code:
                    row_errors.append("ledger_code must belong to only one subgroup_code")
                if ledger_code in seen_ledger_codes:
                    row_errors.append("duplicate ledger_code in upload")
                ledger_subgroup_by_code[ledger_code] = subgroup_code
                seen_ledger_codes.add(ledger_code)

            if row_errors:
                errors.append(
                    {
                        "row_number": row_number,
                        "errors": row_errors,
                    }
                )

        return errors

    async def validate_only(
        self,
        *,
        actor_tenant_id: uuid.UUID,
        file_name: str,
        file_bytes: bytes,
        admitted_airlock_item_id: uuid.UUID,
        airlock_source_type: str,
    ) -> dict[str, Any]:
        await AirlockAdmissionService().assert_admitted(
            self._session,
            tenant_id=actor_tenant_id,
            item_id=admitted_airlock_item_id,
            source_type=airlock_source_type,
        )
        try:
            rows = self.parse_file(file_name=file_name, file_bytes=file_bytes)
        except ValidationError:
            normalized_df = self.parse_normalized_dataframe(
                file_name=file_name,
                file_bytes=file_bytes,
            )
            return {
                "total_rows": int(len(normalized_df.index)),
                "valid_rows": int(len(normalized_df.index)),
                "invalid_rows": 0,
                "errors": [],
            }
        errors = self.validate_structure(rows)
        return {
            "total_rows": len(rows),
            "valid_rows": len(rows) - len(errors),
            "invalid_rows": len(errors),
            "errors": errors,
        }

    @staticmethod
    def _normalize_match_key(value: str) -> str:
        return " ".join(str(value or "").strip().lower().split())

    @classmethod
    def _derive_account_code(cls, account_name: str, *, used_codes: set[str]) -> str:
        base = re.sub(r"[^A-Z0-9]+", "_", str(account_name or "").upper()).strip("_")
        base = base[:42] or f"TB_{uuid.uuid4().hex[:8].upper()}"
        candidate = base
        suffix = 1
        while candidate in used_codes:
            candidate = f"{base[:38]}_{suffix:02d}"[:50]
            suffix += 1
        return candidate[:50]

    @staticmethod
    def _hash_normalized_rows(rows: list[dict[str, Any]]) -> str:
        payload = json.dumps(rows, sort_keys=True, separators=(",", ":"))
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    async def _find_duplicate_flexible_batch(
        self,
        *,
        tenant_id: uuid.UUID,
        template_id: uuid.UUID,
        payload_hash: str,
    ) -> CoaUploadBatch | None:
        rows = (
            await self._session.execute(
                select(CoaUploadBatch)
                .where(
                    CoaUploadBatch.tenant_id == tenant_id,
                    CoaUploadBatch.template_id == template_id,
                )
                .order_by(CoaUploadBatch.created_at.desc())
                .limit(50)
            )
        ).scalars().all()
        for row in rows:
            metadata = dict(row.error_log or {})
            if metadata.get("upload_kind") != _FLEXIBLE_UPLOAD_KIND:
                continue
            if str(metadata.get("normalized_hash") or "") == payload_hash:
                return row
        return None

    async def _build_flexible_activation_plan(
        self,
        *,
        tenant_id: uuid.UUID,
        template_id: uuid.UUID,
        normalized_rows: list[dict[str, Any]],
    ) -> dict[str, Any]:
        tenant_accounts = (
            await self._session.execute(
                select(TenantCoaAccount).where(
                    TenantCoaAccount.tenant_id == tenant_id,
                    TenantCoaAccount.is_active.is_(True),
                )
            )
        ).scalars().all()
        source_ledgers = (
            await self._session.execute(
                select(CoaLedgerAccount).where(
                    CoaLedgerAccount.industry_template_id == template_id,
                    CoaLedgerAccount.is_active.is_(True),
                    or_(
                        CoaLedgerAccount.tenant_id == tenant_id,
                        CoaLedgerAccount.tenant_id.is_(None),
                    ),
                )
            )
        ).scalars().all()

        tenant_by_code = {str(row.account_code or "").strip().upper(): row for row in tenant_accounts}
        tenant_by_name = {
            self._normalize_match_key(row.display_name): row for row in tenant_accounts if row.display_name
        }
        ledger_by_code = {str(row.code or "").strip().upper(): row for row in source_ledgers}
        ledger_by_name = {
            self._normalize_match_key(row.name): row for row in source_ledgers if row.name
        }

        used_codes = set(tenant_by_code.keys()) | set(ledger_by_code.keys())
        plan: list[dict[str, Any]] = []
        mapped_count = 0
        auto_create_count = 0
        review_count = 0

        for row_number, row in enumerate(normalized_rows, start=1):
            account = str(row.get("account") or "").strip()
            normalized_name = self._normalize_match_key(account)
            exact_code = str(account).strip().upper()
            existing = tenant_by_code.get(exact_code) or tenant_by_name.get(normalized_name)
            if existing is not None:
                mapped_count += 1
                plan.append(
                    self._build_flexible_plan_item(
                        row_number=row_number,
                        row=row,
                        status="mapped_existing",
                        tenant_coa_account_id=str(existing.id),
                        account_code=existing.account_code,
                    )
                )
                continue

            ledger = ledger_by_code.get(exact_code) or ledger_by_name.get(normalized_name)
            if ledger is not None:
                auto_create_count += 1
                used_codes.add(ledger.code)
                plan.append(
                    self._build_flexible_plan_item(
                        row_number=row_number,
                        row=row,
                        status="auto_create",
                        ledger_account_id=str(ledger.id),
                        parent_subgroup_id=str(ledger.account_subgroup_id),
                        account_code=ledger.code,
                        display_name=ledger.name,
                    )
                )
                continue

            review_count += 1
            generated_code = self._derive_account_code(account, used_codes=used_codes)
            used_codes.add(generated_code)
            plan.append(
                self._build_flexible_plan_item(
                    row_number=row_number,
                    row=row,
                    status="review",
                    account_code=generated_code,
                )
            )

        return {
            "upload_kind": _FLEXIBLE_UPLOAD_KIND,
            "normalized_rows": normalized_rows,
            "normalized_hash": self._hash_normalized_rows(normalized_rows),
            "activation_plan": plan,
            "activation_summary": {
                "mapped_existing": mapped_count,
                "auto_create": auto_create_count,
                "needs_review": review_count,
                "total_rows": len(normalized_rows),
            },
            "errors": [],
            "total_rows": len(normalized_rows),
            "valid_rows": len(normalized_rows),
            "invalid_rows": 0,
        }

    async def _apply_flexible_batch(
        self,
        *,
        batch: CoaUploadBatch,
        actor_tenant_id: uuid.UUID,
    ) -> dict[str, Any]:
        if batch.tenant_id is None:
            raise ValidationError("Flexible activation requires tenant-scoped batch")
        if batch.tenant_id != actor_tenant_id:
            raise AuthorizationError("Cannot apply another tenant's CoA batch")

        metadata = dict(batch.error_log or {})
        if metadata.get("upload_kind") != _FLEXIBLE_UPLOAD_KIND:
            raise ValidationError("Batch is not a flexible trial balance activation batch")
        if metadata.get("activation_applied") is True:
            summary = dict(metadata.get("activation_summary") or {})
            return {
                "batch_id": str(batch.id),
                "applied_rows": int(summary.get("mapped_existing", 0)) + int(summary.get("auto_create", 0)),
                "template_id": str(batch.template_id),
                "source_type": batch.source_type.value,
            }

        tenant_service = TenantCoaService(self._session)
        if batch.template_id is None:
            raise ValidationError("Flexible activation batch missing template_id")
        if not await tenant_service.has_tenant_accounts(batch.tenant_id):
            await tenant_service.initialise_tenant_coa(batch.tenant_id, batch.template_id)

        existing_accounts = (
            await self._session.execute(
                select(TenantCoaAccount).where(
                    TenantCoaAccount.tenant_id == batch.tenant_id,
                )
            )
        ).scalars().all()
        existing_by_code = {row.account_code: row for row in existing_accounts}
        created = 0
        for item in list(metadata.get("activation_plan") or []):
            if item.get("status") != "auto_create":
                continue
            account_code = str(item.get("account_code") or "").strip().upper()
            if not account_code or account_code in existing_by_code:
                continue
            ledger_id = uuid.UUID(str(item["ledger_account_id"]))
            parent_subgroup_id = uuid.UUID(str(item["parent_subgroup_id"]))
            row = TenantCoaAccount(
                tenant_id=batch.tenant_id,
                ledger_account_id=ledger_id,
                parent_subgroup_id=parent_subgroup_id,
                account_code=account_code,
                display_name=str(item.get("display_name") or item.get("account") or account_code),
                is_custom=False,
                is_active=True,
            )
            self._session.add(row)
            await self._session.flush()
            existing_by_code[account_code] = row
            created += 1

        summary = dict(metadata.get("activation_summary") or {})
        summary["auto_created_applied"] = created
        metadata["activation_summary"] = summary
        metadata["activation_applied"] = True
        batch.error_log = metadata
        batch.processed_at = datetime.now(UTC)
        batch.upload_status = CoaUploadStatus.SUCCESS
        await self._session.flush()
        return {
            "batch_id": str(batch.id),
            "applied_rows": int(summary.get("mapped_existing", 0)) + created,
            "template_id": str(batch.template_id),
            "source_type": batch.source_type.value,
        }

    async def upload(
        self,
        *,
        actor_id: uuid.UUID,
        actor_tenant_id: uuid.UUID,
        tenant_id: uuid.UUID | None,
        template_id: uuid.UUID,
        source_type: CoaSourceType,
        upload_mode: CoaUploadMode,
        file_name: str,
        file_bytes: bytes,
        admitted_airlock_item_id: uuid.UUID,
        airlock_source_type: str,
    ) -> dict[str, Any]:
        await AirlockAdmissionService().assert_admitted(
            self._session,
            tenant_id=actor_tenant_id,
            item_id=admitted_airlock_item_id,
            source_type=airlock_source_type,
        )
        template = (
            await self._session.execute(
                select(CoaIndustryTemplate).where(CoaIndustryTemplate.id == template_id)
            )
        ).scalar_one_or_none()
        if template is None:
            raise NotFoundError("Industry template not found")

        normalized_df = self.parse_normalized_dataframe(
            file_name=file_name,
            file_bytes=file_bytes,
        )
        is_canonical_upload = set(_REQUIRED_COLUMNS).issubset(set(normalized_df.columns))
        flexible_rows = (
            normalized_df[["account", "debit", "credit"]].to_dict(orient="records")
            if not is_canonical_upload
            else []
        )
        if not is_canonical_upload:
            if tenant_id is None:
                raise ValidationError(
                    "Flexible trial balance activation is only supported for tenant-scoped uploads"
                )
            normalized_hash = self._hash_normalized_rows(flexible_rows)
            duplicate = await self._find_duplicate_flexible_batch(
                tenant_id=tenant_id,
                template_id=template_id,
                payload_hash=normalized_hash,
            )
            if duplicate is not None:
                metadata = dict(duplicate.error_log or {})
                summary = dict(metadata.get("activation_summary") or {})
                return {
                    "batch_id": str(duplicate.id),
                    "upload_status": duplicate.upload_status.value,
                    "total_rows": int(metadata.get("total_rows", 0)),
                    "valid_rows": int(metadata.get("valid_rows", 0)),
                    "invalid_rows": int(metadata.get("invalid_rows", 0)),
                    "errors": list(metadata.get("errors", [])),
                    "upload_kind": str(metadata.get("upload_kind") or _FLEXIBLE_UPLOAD_KIND),
                    "activation_summary": summary,
                    "requires_review": int(summary.get("needs_review", 0) or 0) > 0,
                    "idempotent_replay": True,
                }

        batch = CoaUploadBatch(
            tenant_id=tenant_id,
            template_id=template_id,
            source_type=source_type,
            upload_mode=upload_mode,
            file_name=file_name,
            upload_status=CoaUploadStatus.PROCESSING,
            created_by=actor_id,
        )
        self._session.add(batch)
        await self._session.flush()

        if not is_canonical_upload:
            metadata = await self._build_flexible_activation_plan(
                tenant_id=tenant_id,
                template_id=template_id,
                normalized_rows=flexible_rows,
            )
            batch.upload_status = CoaUploadStatus.SUCCESS
            batch.error_log = self._json_safe_value(metadata)
            batch.processed_at = datetime.now(UTC)
            await self._session.flush()
            return {
                "batch_id": str(batch.id),
                "upload_status": batch.upload_status.value,
                "total_rows": int(metadata["total_rows"]),
                "valid_rows": int(metadata["valid_rows"]),
                "invalid_rows": int(metadata["invalid_rows"]),
                "errors": list(metadata["errors"]),
                "upload_kind": _FLEXIBLE_UPLOAD_KIND,
                "activation_summary": dict(metadata["activation_summary"]),
                "requires_review": int(metadata["activation_summary"].get("needs_review", 0) or 0) > 0,
                "idempotent_replay": False,
            }

        rows = self.parse_file(file_name=file_name, file_bytes=file_bytes)
        errors = self.validate_structure(rows)
        error_map = {int(item["row_number"]): list(item["errors"]) for item in errors}

        if upload_mode != CoaUploadMode.VALIDATE_ONLY:
            staging_payload = []
            for row in rows:
                row_number = int(row["row_number"])
                staging_payload.append(
                    {
                        "batch_id": batch.id,
                        "tenant_id": tenant_id,
                        "template_id": template_id,
                        "row_number": row_number,
                        "group_code": row["group_code"],
                        "group_name": row["group_name"],
                        "subgroup_code": row["subgroup_code"],
                        "subgroup_name": row["subgroup_name"],
                        "ledger_code": row["ledger_code"],
                        "ledger_name": row["ledger_name"],
                        "ledger_type": row["ledger_type"],
                        "is_control_account": bool(row["is_control_account"]),
                        "validation_errors": [{"message": err} for err in error_map.get(row_number, [])] or None,
                        "is_valid": row_number not in error_map,
                    }
                )
            if staging_payload:
                await self._session.execute(insert(CoaUploadStagingRow).values(staging_payload))

        batch.upload_status = CoaUploadStatus.FAILED if errors else CoaUploadStatus.SUCCESS
        batch.error_log = {
            "errors": errors,
            "total_rows": len(rows),
            "valid_rows": len(rows) - len(errors),
            "invalid_rows": len(errors),
        }
        batch.processed_at = datetime.now(UTC)
        await self._session.flush()

        return {
            "batch_id": str(batch.id),
            "upload_status": batch.upload_status.value,
            "total_rows": len(rows),
            "valid_rows": len(rows) - len(errors),
            "invalid_rows": len(errors),
            "errors": errors,
            "upload_kind": None,
            "activation_summary": None,
            "requires_review": False,
            "idempotent_replay": False,
        }

    async def _next_version(
        self,
        *,
        template_id: uuid.UUID,
        tenant_id: uuid.UUID | None,
        source_type: CoaSourceType,
    ) -> int:
        max_version = (
            await self._session.execute(
                select(func.max(CoaLedgerAccount.version)).where(
                    CoaLedgerAccount.industry_template_id == template_id,
                    CoaLedgerAccount.source_type == source_type,
                    CoaLedgerAccount.tenant_id.is_(tenant_id)
                    if tenant_id is None
                    else CoaLedgerAccount.tenant_id == tenant_id,
                )
            )
        ).scalar_one_or_none()
        return int(max_version or 0) + 1

    async def _upsert_groups(
        self,
        *,
        template_id: uuid.UUID,
        rows: list[CoaUploadStagingRow],
    ) -> dict[str, CoaAccountGroup]:
        group_pairs = {(row.group_code, row.group_name) for row in rows}
        group_codes = [item[0] for item in group_pairs]

        existing = (
            await self._session.execute(
                select(CoaAccountGroup).where(
                    CoaAccountGroup.industry_template_id == template_id,
                    CoaAccountGroup.code.in_(group_codes),
                )
            )
        ).scalars().all()
        by_code = {row.code: row for row in existing}

        max_sort = (
            await self._session.execute(
                select(func.max(CoaAccountGroup.sort_order)).where(
                    CoaAccountGroup.industry_template_id == template_id
                )
            )
        ).scalar_one_or_none() or 0

        for code, name in sorted(group_pairs):
            row = by_code.get(code)
            if row is None:
                max_sort += 1
                row = CoaAccountGroup(
                    industry_template_id=template_id,
                    fs_subline_id=None,
                    code=code,
                    name=name,
                    sort_order=int(max_sort),
                    is_active=True,
                )
                self._session.add(row)
                by_code[code] = row
            else:
                row.name = name
                row.is_active = True

        await self._session.flush()
        return by_code

    async def _upsert_subgroups(
        self,
        *,
        groups_by_code: dict[str, CoaAccountGroup],
        rows: list[CoaUploadStagingRow],
    ) -> dict[str, CoaAccountSubgroup]:
        subgroup_pairs = {
            (row.subgroup_code, row.subgroup_name, row.group_code)
            for row in rows
        }

        group_ids = [group.id for group in groups_by_code.values()]
        existing = (
            await self._session.execute(
                select(CoaAccountSubgroup).where(CoaAccountSubgroup.account_group_id.in_(group_ids))
            )
        ).scalars().all()

        by_key = {(row.account_group_id, row.code): row for row in existing}
        max_sort_by_group: dict[uuid.UUID, int] = defaultdict(int)
        for row in existing:
            max_sort_by_group[row.account_group_id] = max(max_sort_by_group[row.account_group_id], row.sort_order)

        by_code: dict[str, CoaAccountSubgroup] = {}
        for subgroup_code, subgroup_name, group_code in sorted(subgroup_pairs):
            group = groups_by_code[group_code]
            key = (group.id, subgroup_code)
            row = by_key.get(key)
            if row is None:
                next_sort = max_sort_by_group[group.id] + 1
                max_sort_by_group[group.id] = next_sort
                row = CoaAccountSubgroup(
                    account_group_id=group.id,
                    code=subgroup_code,
                    name=subgroup_name,
                    sort_order=next_sort,
                    is_active=True,
                )
                self._session.add(row)
                by_key[key] = row
            else:
                row.name = subgroup_name
                row.is_active = True
            by_code[subgroup_code] = row

        await self._session.flush()
        return by_code

    @staticmethod
    def _normal_balance_from_ledger_type(ledger_type: str) -> str:
        if ledger_type in {"ASSET", "EXPENSE"}:
            return "DEBIT"
        return "CREDIT"

    @staticmethod
    def _bspl_from_ledger_type(ledger_type: str) -> str:
        if ledger_type == "INCOME":
            return "REVENUE"
        return ledger_type

    async def _upsert_ledgers(
        self,
        *,
        batch: CoaUploadBatch,
        rows: list[CoaUploadStagingRow],
        subgroups_by_code: dict[str, CoaAccountSubgroup],
    ) -> int:
        if batch.template_id is None:
            raise ValidationError("Batch missing template_id")

        version = await self._next_version(
            template_id=batch.template_id,
            tenant_id=batch.tenant_id,
            source_type=batch.source_type,
        )

        if batch.upload_mode == CoaUploadMode.REPLACE:
            await self._session.execute(
                update(CoaLedgerAccount)
                .where(
                    CoaLedgerAccount.industry_template_id == batch.template_id,
                    CoaLedgerAccount.source_type == batch.source_type,
                    CoaLedgerAccount.tenant_id.is_(batch.tenant_id)
                    if batch.tenant_id is None
                    else CoaLedgerAccount.tenant_id == batch.tenant_id,
                    CoaLedgerAccount.is_active.is_(True),
                )
                .values(is_active=False)
            )

        payload = []
        for idx, row in enumerate(rows, start=1):
            subgroup = subgroups_by_code[row.subgroup_code]
            payload.append(
                {
                    "account_subgroup_id": subgroup.id,
                    "industry_template_id": batch.template_id,
                    "tenant_id": batch.tenant_id,
                    "code": row.ledger_code,
                    "name": row.ledger_name,
                    "description": f"Uploaded via CoA batch {batch.id}",
                    "source_type": batch.source_type,
                    "version": version,
                    "created_by": batch.created_by,
                    "normal_balance": self._normal_balance_from_ledger_type(row.ledger_type),
                    "cash_flow_tag": None,
                    "cash_flow_method": None,
                    "bs_pl_flag": self._bspl_from_ledger_type(row.ledger_type),
                    "asset_liability_class": None,
                    "is_monetary": False,
                    "is_related_party": False,
                    "is_tax_deductible": True,
                    "is_control_account": row.is_control_account,
                    "notes_reference": None,
                    "is_active": True,
                    "sort_order": idx,
                }
            )

        if not payload:
            return 0

        if batch.tenant_id is None:
            await self._session.execute(insert(CoaLedgerAccount).values(payload))
            return len(payload)

        stmt = insert(CoaLedgerAccount).values(payload)
        excluded = stmt.excluded
        stmt = stmt.on_conflict_do_update(
            constraint="uq_coa_ledger_accounts_tenant_code",
            set_={
                "account_subgroup_id": excluded.account_subgroup_id,
                "industry_template_id": excluded.industry_template_id,
                "name": excluded.name,
                "description": excluded.description,
                "source_type": excluded.source_type,
                "version": excluded.version,
                "created_by": excluded.created_by,
                "normal_balance": excluded.normal_balance,
                "cash_flow_tag": excluded.cash_flow_tag,
                "cash_flow_method": excluded.cash_flow_method,
                "bs_pl_flag": excluded.bs_pl_flag,
                "asset_liability_class": excluded.asset_liability_class,
                "is_monetary": excluded.is_monetary,
                "is_related_party": excluded.is_related_party,
                "is_tax_deductible": excluded.is_tax_deductible,
                "is_control_account": excluded.is_control_account,
                "notes_reference": excluded.notes_reference,
                "is_active": True,
                "sort_order": excluded.sort_order,
                "updated_at": func.now(),
            },
        )
        await self._session.execute(stmt)
        return len(payload)

    async def apply_batch(
        self,
        *,
        batch_id: uuid.UUID,
        actor_tenant_id: uuid.UUID,
        is_platform_admin: bool,
    ) -> dict[str, Any]:
        batch = (
            await self._session.execute(
                select(CoaUploadBatch).where(CoaUploadBatch.id == batch_id)
            )
        ).scalar_one_or_none()
        if batch is None:
            raise NotFoundError("CoA upload batch not found")

        if batch.upload_mode == CoaUploadMode.VALIDATE_ONLY:
            raise ValidationError("VALIDATE_ONLY batch cannot be applied")
        if batch.upload_status != CoaUploadStatus.SUCCESS:
            raise ValidationError("Only SUCCESS batches can be applied")

        if batch.source_type == CoaSourceType.ADMIN_TEMPLATE and not is_platform_admin:
            raise AuthorizationError("Only platform admins can apply ADMIN_TEMPLATE batches")

        if batch.source_type == CoaSourceType.TENANT_CUSTOM:
            if batch.tenant_id is None:
                raise ValidationError("Tenant-scoped batch missing tenant_id")
            if batch.tenant_id != actor_tenant_id:
                raise AuthorizationError("Cannot apply another tenant's CoA batch")

        metadata = dict(batch.error_log or {})
        if metadata.get("upload_kind") == _FLEXIBLE_UPLOAD_KIND:
            return await self._apply_flexible_batch(
                batch=batch,
                actor_tenant_id=actor_tenant_id,
            )

        rows = (
            await self._session.execute(
                select(CoaUploadStagingRow)
                .where(
                    CoaUploadStagingRow.batch_id == batch.id,
                    CoaUploadStagingRow.is_valid.is_(True),
                )
                .order_by(CoaUploadStagingRow.row_number.asc())
            )
        ).scalars().all()
        if not rows:
            raise ValidationError("No valid staging rows available for apply")

        groups_by_code = await self._upsert_groups(template_id=batch.template_id, rows=rows)
        subgroups_by_code = await self._upsert_subgroups(groups_by_code=groups_by_code, rows=rows)
        inserted = await self._upsert_ledgers(batch=batch, rows=rows, subgroups_by_code=subgroups_by_code)

        if batch.template_id is None:
            raise ValidationError("Batch missing template_id")

        group_count = int(
            (
                await self._session.execute(
                    select(func.count())
                    .select_from(CoaAccountGroup)
                    .where(
                        CoaAccountGroup.industry_template_id == batch.template_id,
                        CoaAccountGroup.is_active.is_(True),
                    )
                )
            ).scalar_one()
        )
        subgroup_count = int(
            (
                await self._session.execute(
                    select(func.count())
                    .select_from(CoaAccountSubgroup)
                    .join(
                        CoaAccountGroup,
                        CoaAccountSubgroup.account_group_id == CoaAccountGroup.id,
                    )
                    .where(
                        CoaAccountGroup.industry_template_id == batch.template_id,
                        CoaAccountGroup.is_active.is_(True),
                        CoaAccountSubgroup.is_active.is_(True),
                    )
                )
            ).scalar_one()
        )
        ledger_count_stmt = (
            select(func.count())
            .select_from(CoaLedgerAccount)
            .where(
                CoaLedgerAccount.industry_template_id == batch.template_id,
                CoaLedgerAccount.source_type == batch.source_type,
                CoaLedgerAccount.is_active.is_(True),
            )
        )
        if batch.tenant_id is None:
            ledger_count_stmt = ledger_count_stmt.where(CoaLedgerAccount.tenant_id.is_(None))
        else:
            ledger_count_stmt = ledger_count_stmt.where(CoaLedgerAccount.tenant_id == batch.tenant_id)
        ledger_count = int((await self._session.execute(ledger_count_stmt)).scalar_one())

        if group_count <= 0:
            raise ValidationError("CoA apply validation failed: group_count must be > 0")
        if subgroup_count <= 0:
            raise ValidationError("CoA apply validation failed: subgroup_count must be > 0")
        if ledger_count <= 0:
            raise ValidationError("CoA apply validation failed: ledger_count must be > 0")

        log.info(
            "coa_apply_completed upload_id=%s tenant_id=%s ledger_count_inserted=%s",
            batch.id,
            batch.tenant_id,
            inserted,
        )

        batch.processed_at = datetime.now(UTC)
        batch.upload_status = CoaUploadStatus.SUCCESS
        await self._session.flush()

        return {
            "batch_id": str(batch.id),
            "applied_rows": inserted,
            "template_id": str(batch.template_id),
            "source_type": batch.source_type.value,
        }
