from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from financeops.modules.board_pack_generator.domain.pack_definition import AssembledPack


def get_weasyprint_mode() -> str:
    """
    Returns 'real' if WeasyPrint can render with system libraries,
    'stub' if running in stub/mock mode (Windows dev),
    'broken' if WeasyPrint is installed but system libs are missing.
    """
    try:
        import weasyprint
    except OSError:
        return "broken"
    except Exception:
        return "stub"

    try:
        weasyprint.HTML(string="<p>smoke</p>").write_pdf()
        return "real"
    except OSError:
        return "broken"
    except Exception:
        return "stub"


def assert_weasyprint_available() -> None:
    """
    Raises RuntimeError if WeasyPrint is in broken mode.
    Call this at board pack generation time to fail loudly
    rather than returning a corrupt PDF.
    """
    mode = get_weasyprint_mode()
    if mode == "broken":
        raise RuntimeError(
            "WeasyPrint is installed but system libraries are missing "
            "(libgobject, Cairo, Pango). Install system dependencies "
            "or use the Linux deployment container."
        )


def _format_decimal(value: Any) -> Any:
    if isinstance(value, Decimal):
        text = format(value, "f")
        sign = ""
        if text.startswith("-"):
            sign = "-"
            text = text[1:]
        if "." in text:
            int_part, frac_part = text.split(".", 1)
            int_formatted = f"{int(int_part or '0'):,}"
            frac_trimmed = frac_part.rstrip("0")
            if frac_trimmed:
                return f"{sign}{int_formatted}.{frac_trimmed}"
            return f"{sign}{int_formatted}"
        return f"{sign}{int(text or '0'):,}"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _format_decimal(child) for key, child in value.items()}
    if isinstance(value, list):
        return [_format_decimal(item) for item in value]
    if isinstance(value, tuple):
        return [_format_decimal(item) for item in value]
    return value


def _to_cell_value(value: Any) -> str:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if value is None:
        return ""
    return str(value)


def _flatten_dict_rows(payload: dict[str, Any], prefix: str = "") -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    for key, value in payload.items():
        current_key = f"{prefix}.{key}" if prefix else str(key)
        if isinstance(value, dict):
            rows.extend(_flatten_dict_rows(value, prefix=current_key))
            continue
        rows.append((current_key, value))
    return rows


def _as_table(payload: Any) -> tuple[list[str], list[list[Any]]]:
    if isinstance(payload, list) and payload and all(isinstance(item, dict) for item in payload):
        headers: list[str] = []
        for item in payload:
            for key in item:
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)
        rows: list[list[Any]] = []
        for item in payload:
            row = [_to_cell_value(item.get(header)) for header in headers]
            rows.append(row)
        return headers, rows

    if isinstance(payload, dict):
        rows = _flatten_dict_rows(payload)
        return ["Key", "Value"], [[key, _to_cell_value(value)] for key, value in rows]

    if isinstance(payload, list):
        return ["Index", "Value"], [[str(index), _to_cell_value(item)] for index, item in enumerate(payload, start=1)]

    return ["Value"], [[_to_cell_value(payload)]]


def _safe_sheet_title(title: str, existing_titles: set[str]) -> str:
    invalid = {'\\', '/', '*', '[', ']', ':', '?'}
    cleaned = "".join(char if char not in invalid else "_" for char in title).strip() or "Section"
    base = cleaned[:31]
    candidate = base
    suffix = 1
    while candidate in existing_titles:
        suffix_text = f"_{suffix}"
        candidate = f"{base[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    existing_titles.add(candidate)
    return candidate


class BoardPackExportService:
    def __init__(self) -> None:
        self._templates_dir = Path(__file__).resolve().parent.parent / "templates"

    def export_pdf(
        self,
        pack: AssembledPack,
        pack_name: str,
        generated_at: datetime,
    ) -> tuple[bytes, str]:
        """
        Render AssembledPack to PDF bytes using WeasyPrint + Jinja2.
        Returns (pdf_bytes, suggested_filename).
        Filename format: board_pack_{period_start}_{period_end}.pdf
        All Decimal values must be formatted as strings, never binary inexact numeric types.
        """
        assert_weasyprint_available()

        from jinja2 import Environment, FileSystemLoader, select_autoescape
        from weasyprint import HTML

        jinja_env = Environment(
            loader=FileSystemLoader(str(self._templates_dir)),
            autoescape=select_autoescape(["html", "xml"]),
        )

        ordered_sections = sorted(pack.sections, key=lambda section: section.section_order)
        sections_payload = [
            {
                "section_type": section.section_type.value,
                "section_order": section.section_order,
                "title": section.title,
                "data_snapshot": _format_decimal(section.data_snapshot),
                "section_hash": section.section_hash,
            }
            for section in ordered_sections
        ]

        template = jinja_env.get_template("board_pack.html")
        rendered_html = template.render(
            pack_name=pack_name,
            period_start=pack.period_start.isoformat(),
            period_end=pack.period_end.isoformat(),
            generated_at=generated_at.isoformat(),
            chain_hash=pack.chain_hash,
            sections=sections_payload,
        )

        pdf_bytes = HTML(string=rendered_html, base_url=str(self._templates_dir)).write_pdf()
        if not pdf_bytes:
            raise RuntimeError("WeasyPrint returned empty PDF output")
        suggested_filename = f"board_pack_{pack.period_start.isoformat()}_{pack.period_end.isoformat()}.pdf"
        return pdf_bytes, suggested_filename

    def export_excel(
        self,
        pack: AssembledPack,
        pack_name: str,
        generated_at: datetime,
    ) -> tuple[bytes, str]:
        """
        Render AssembledPack to Excel bytes using openpyxl.
        Returns (excel_bytes, suggested_filename).
        Filename format: board_pack_{period_start}_{period_end}.xlsx

        Workbook structure:
        - Sheet 1: "Cover" - pack name, period, generated timestamp
        - One sheet per RenderedSection named after the section title
          (truncated to 31 chars - Excel sheet name limit)
        - Each sheet: header row (bold, light blue fill) +
          data rows from the section's payload
        - Decimal values rendered as strings in cells - never binary inexact numeric types
        - Column widths auto-sized (min 12, max 50)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        cover_sheet = workbook.active
        cover_sheet.title = "Cover"
        cover_sheet["A1"] = "Board Pack"
        cover_sheet["B1"] = pack_name
        cover_sheet["A2"] = "Period Start"
        cover_sheet["B2"] = pack.period_start.isoformat()
        cover_sheet["A3"] = "Period End"
        cover_sheet["B3"] = pack.period_end.isoformat()
        cover_sheet["A4"] = "Generated At"
        cover_sheet["B4"] = generated_at.isoformat()
        cover_sheet["A5"] = "Chain Hash"
        cover_sheet["B5"] = pack.chain_hash
        for cell in ("A1", "A2", "A3", "A4", "A5"):
            cover_sheet[cell].font = Font(bold=True)
        cover_sheet.column_dimensions["A"].width = 18
        cover_sheet.column_dimensions["B"].width = 50

        header_fill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
        ordered_sections = sorted(pack.sections, key=lambda section: section.section_order)
        used_titles = {"Cover"}

        for section in ordered_sections:
            title = _safe_sheet_title(section.title, used_titles)
            sheet = workbook.create_sheet(title=title)
            formatted_payload = _format_decimal(section.data_snapshot)
            headers, rows = _as_table(formatted_payload)

            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            for row in rows:
                sheet.append([_to_cell_value(value) for value in row])

            sheet.freeze_panes = "A2"

            for column_index, column_cells in enumerate(sheet.columns, start=1):
                max_length = 0
                for cell in column_cells:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                width = max(12, min(50, max_length + 2))
                sheet.column_dimensions[get_column_letter(column_index)].width = width

        output = BytesIO()
        workbook.save(output)
        excel_bytes = output.getvalue()
        suggested_filename = f"board_pack_{pack.period_start.isoformat()}_{pack.period_end.isoformat()}.xlsx"
        return excel_bytes, suggested_filename


__all__ = [
    "BoardPackExportService",
    "assert_weasyprint_available",
    "get_weasyprint_mode",
]
