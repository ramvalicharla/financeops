from __future__ import annotations

import csv
import json
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any


def _format_decimal(value: Any) -> Any:
    if isinstance(value, Decimal):
        text = format(value, "f")
        sign = ""
        if text.startswith("-"):
            sign = "-"
            text = text[1:]
        if "." in text:
            int_part, frac_part = text.split(".", 1)
            int_formatted = f"{int(int_part or '0'):,}"
            frac_trimmed = frac_part.rstrip("0")
            if frac_trimmed:
                return f"{sign}{int_formatted}.{frac_trimmed}"
            return f"{sign}{int_formatted}"
        return f"{sign}{int(text or '0'):,}"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, dict):
        return {str(key): _format_decimal(child) for key, child in value.items()}
    if isinstance(value, list):
        return [_format_decimal(item) for item in value]
    if isinstance(value, tuple):
        return [_format_decimal(item) for item in value]
    return value


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return str(value)


class ReportExportService:
    def export_csv(
        self,
        rows: list[dict[str, Any]],
        report_name: str,
    ) -> tuple[bytes, str]:
        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)

        output = StringIO(newline="")
        writer = csv.writer(output)
        if headers:
            writer.writerow(headers)
        for row in rows:
            formatted = _format_decimal(row)
            writer.writerow([_cell(formatted.get(header)) for header in headers])

        csv_bytes = output.getvalue().encode("utf-8")
        filename = f"{report_name.strip().lower().replace(' ', '_') or 'custom_report'}.csv"
        return csv_bytes, filename

    def export_excel(
        self,
        rows: list[dict[str, Any]],
        report_name: str,
    ) -> tuple[bytes, str]:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"

        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)

        if headers:
            worksheet.append(headers)
            header_fill = PatternFill(fill_type="solid", start_color="D9E1F2", end_color="D9E1F2")
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

        for row in rows:
            formatted = _format_decimal(row)
            worksheet.append([_cell(formatted.get(header)) for header in headers])

        for idx, column_cells in enumerate(worksheet.columns, start=1):
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in column_cells)
            worksheet.column_dimensions[get_column_letter(idx)].width = max(12, min(50, max_len + 2))

        payload = BytesIO()
        workbook.save(payload)
        filename = f"{report_name.strip().lower().replace(' ', '_') or 'custom_report'}.xlsx"
        return payload.getvalue(), filename

    def export_pdf(
        self,
        rows: list[dict[str, Any]],
        report_name: str,
        generated_at: datetime,
    ) -> tuple[bytes, str]:
        from jinja2 import BaseLoader, Environment, select_autoescape
        from weasyprint import HTML

        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in headers:
                    headers.append(key_text)

        formatted_rows: list[list[str]] = []
        for row in rows:
            formatted = _format_decimal(row)
            formatted_rows.append([_cell(formatted.get(header)) for header in headers])

        template = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <style>
    body { font-family: Arial, sans-serif; color: #111; font-size: 12px; margin: 24px; }
    h1 { font-size: 18px; margin: 0 0 8px; }
    .meta { margin-bottom: 16px; }
    table { width: 100%; border-collapse: collapse; table-layout: fixed; }
    th, td { border: 1px solid #d1d5db; padding: 6px 8px; word-wrap: break-word; }
    th { background: #e5e7eb; text-align: left; }
    .footer { margin-top: 16px; font-size: 10px; color: #555; }
  </style>
</head>
<body>
  <h1>{{ report_name }}</h1>
  <div class="meta">Generated at: {{ generated_at }}</div>
  <table>
    <thead>
      <tr>
        {% for header in headers %}
          <th>{{ header }}</th>
        {% endfor %}
      </tr>
    </thead>
    <tbody>
      {% for row in rows %}
        <tr>
          {% for cell in row %}
            <td>{{ cell }}</td>
          {% endfor %}
        </tr>
      {% endfor %}
    </tbody>
  </table>
  <div class="footer">Confidential - FinanceOps</div>
</body>
</html>
""".strip()

        env = Environment(loader=BaseLoader(), autoescape=select_autoescape(["html", "xml"]))
        html = env.from_string(template).render(
            report_name=report_name,
            generated_at=generated_at.isoformat(),
            headers=headers,
            rows=formatted_rows,
        )
        pdf_bytes = HTML(string=html).write_pdf()
        filename = f"{report_name.strip().lower().replace(' ', '_') or 'custom_report'}.pdf"
        return pdf_bytes, filename


__all__ = ["ReportExportService"]

